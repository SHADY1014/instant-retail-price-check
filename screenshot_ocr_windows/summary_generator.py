"""
汇总表生成器 - 分省汇总 + 分地级市汇总（同一Sheet）+ 明细表(含截图，独立Sheet)

从巡查表 xlsx 读取数据和图片，生成汇总报告。
使用巡查表模板的样式（深蓝标题、蓝表头、居中数据）。
图片使用 WPS DISPIMG 格式嵌入单元格。

Sheet1 "汇总表"：分省汇总 + 分地级市汇总，上下排列2个区块
Sheet2 "明细表"：逐条记录含判定结果 + 截图
"""

import io
import os
import re
import shutil
import uuid
import zipfile
from collections import defaultdict
from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# =========================================================
# 合格标准
# =========================================================
QUALIFICATION_RULES = [
    ("漓泉1998", "500ml*12瓶", 74.99),
    ("漓泉1998", "500ml*12听", 74.99),
    ("漓泉1998", "500ml*9听", 45),
    ("漓泉1998", "500ml*9+3听", 45),
    ("燕京U8", "500ml*12瓶", 60),
    ("燕京U8", "500ml*12听", 60),
]

# 1998 产品的所有名称变体（OCR统一为"漓泉1998啤酒"，人工可能改为"铂金1998啤酒"/"特渠1998啤酒"）
_1998_KEYWORDS = ("漓泉1998", "铂金1998", "特渠1998")
_1998_12_SPECS = {"500ml*12瓶", "500ml*12听"}
# 1998 12瓶/12听按销售省份执行不同价格口径。
# 未配置省份暂保留旧口径，避免对历史数据做无依据的推断。
_1998_PROVINCE_THRESHOLDS = {
    "广东": (70, 65),
    "广西": (60, 55),
}


def _is_1998_product(product_name):
    """判断是否为1998系列产品（兼容漓泉/铂金/特渠三种命名）"""
    return any(kw in product_name for kw in _1998_KEYWORDS)

# =========================================================
# 城市所属省份反查表
# =========================================================
CITY_TO_PROVINCE = {}
_POOL = {
    "广东": ["广州", "深圳", "珠海", "汕头", "佛山", "韶关", "湛江", "肇庆",
             "江门", "茂名", "惠州", "梅州", "汕尾", "河源", "阳江", "清远",
             "东莞", "中山", "潮州", "揭阳", "云浮"],
    "广西": ["南宁", "柳州", "桂林", "梧州", "北海", "防城港", "钦州", "贵港",
             "玉林", "百色", "贺州", "河池", "来宾", "崇左"],
    "海南": ["海口", "三亚", "三沙", "儋州"],
    "贵州": ["贵阳", "六盘水", "遵义", "安顺", "毕节", "铜仁"],
    "云南": ["昆明", "曲靖", "玉溪", "保山", "昭通", "丽江", "普洱", "临沧"],
}
for prov, cities in _POOL.items():
    for city in cities:
        CITY_TO_PROVINCE[city + "市"] = prov
        CITY_TO_PROVINCE[city] = prov

# =========================================================
# 路径
# =========================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(BASE_DIR, "燕京啤酒即时零售渠道价格巡查表-备份.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# =========================================================
# 样式定义（与巡查表模板一致）
# =========================================================
FONT_NAME = "微软雅黑"

# 标题行（深蓝底主题浅色 1 字，与模板 A1 一致）
STYLE_TITLE = {
    "font": Font(name=FONT_NAME, size=14, bold=True, color=Color(theme=0)),
    "fill": PatternFill(fill_type="solid", fgColor="FF003366"),
    "align": Alignment(horizontal="center", vertical="center"),
}

# 副标题/合格标准行
STYLE_SUBTITLE = {
    "font": Font(name=FONT_NAME, size=9, bold=False, color="FF666666"),
    "fill": PatternFill(fill_type=None),
    "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
}

# 表头行（蓝底主题浅色 1 字，与模板第2行一致）
STYLE_HEADER = {
    "font": Font(name=FONT_NAME, size=11, bold=True, color=Color(theme=0)),
    "fill": PatternFill(fill_type="solid", fgColor="FF0070C0"),
    "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
}

# 数据行
STYLE_DATA = {
    "font": Font(name=FONT_NAME, size=10, bold=False),
    "fill": PatternFill(fill_type=None),
    "align": Alignment(horizontal="center", vertical="center"),
}

# 合格样式（绿底绿字）
STYLE_PASS = {
    "font": Font(name=FONT_NAME, size=10, bold=True, color="FF008000"),
    "fill": PatternFill(fill_type="solid", fgColor="FFC6EFCE"),
    "align": Alignment(horizontal="center", vertical="center"),
}

# 不合格样式（红底红字）
STYLE_FAIL = {
    "font": Font(name=FONT_NAME, size=10, bold=True, color="FFFF0000"),
    "fill": PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
    "align": Alignment(horizontal="center", vertical="center"),
}

# 总计行
STYLE_TOTAL = {
    "font": Font(name=FONT_NAME, size=10, bold=True),
    "fill": PatternFill(fill_type="solid", fgColor="FFD9E1F2"),
    "align": Alignment(horizontal="center", vertical="center"),
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="FFB0B0B0"),
    right=Side(style="thin", color="FFB0B0B0"),
    top=Side(style="thin", color="FFB0B0B0"),
    bottom=Side(style="thin", color="FFB0B0B0"),
)


def _apply_style(cell, style_dict):
    """应用样式到单元格"""
    if "font" in style_dict:
        cell.font = copy(style_dict["font"])
    if "fill" in style_dict:
        cell.fill = copy(style_dict["fill"])
    if "align" in style_dict:
        cell.alignment = copy(style_dict["align"])
    cell.border = copy(_THIN_BORDER)


# =========================================================
# 数据提取工具
# =========================================================

def _extract_spec(product_name):
    """从产品名中提取规格，如 '500ml*12瓶' 或 '500ml*9+3听'"""
    # 支持 "9+3听" 这种组合数量格式
    m = re.search(r'(\d+ml\*\d+\+\d+[瓶听罐])', product_name)
    if m:
        spec = m.group(1)
        spec = spec.replace("罐", "听")
        return spec
    m = re.search(r'(\d+ml\*\d+[瓶听罐])', product_name)
    if m:
        spec = m.group(1)
        spec = spec.replace("罐", "听")
        return spec
    return ""


def _safe_float(val):
    """安全转 float，处理 None/空/¥符号/逗号/全角数字等格式

    Examples:
        _safe_float("75.9")     -> 75.9
        _safe_float("¥75.9")    -> 75.9
        _safe_float("3元")      -> 3.0
        _safe_float("1,234.5")  -> 1234.5
        _safe_float(None)       -> 0.0
        _safe_float("")         -> 0.0
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    # 去掉货币符号、单位
    s = re.sub(r'[¥$￥]', '', s)
    s = re.sub(r'[元块]', '', s)
    # 去掉千分位逗号
    s = s.replace(',', '').replace('，', '')
    # 全角数字转半角
    s = s.translate(str.maketrans('０１２３４５６７８９．', '0123456789.'))
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _get_base_qualification_line(product_name, spec):
    """根据产品名和规格获取省份覆盖前的基础合格线。"""
    # 1998 系列：铂金1998/特渠1998 视同 漓泉1998
    normalized = product_name
    if _is_1998_product(product_name):
        normalized = product_name
        for kw in ("铂金1998", "特渠1998"):
            normalized = normalized.replace(kw, "漓泉1998")
    for kw_product, kw_spec, line in QUALIFICATION_RULES:
        if kw_product in normalized and kw_spec in spec:
            return line
    return None


def _get_qualification_line(product_name, spec, province=""):
    """根据产品、规格和省份获取合格线。"""
    base_line = _get_base_qualification_line(product_name, spec)
    if base_line is None:
        return None
    if _is_1998_product(product_name) and spec in _1998_12_SPECS:
        return _1998_PROVINCE_THRESHOLDS.get(province, (base_line, 70))[0]
    return base_line


def _get_secondary_threshold(product_name, province="", spec=""):
    """获取第二档线：1998按省份，U8为55元。"""
    if _is_1998_product(product_name):
        if spec in _1998_12_SPECS:
            base_line = _get_base_qualification_line(product_name, spec)
            return _1998_PROVINCE_THRESHOLDS.get(province, (base_line, 70))[1]
        return 70
    return 55


def _get_record_thresholds(record):
    """Return ``(qualification_line, secondary_line)`` for one report row."""
    product_name = str(record.get("product_name", ""))
    spec = str(record.get("spec", ""))
    province = str(record.get("province", ""))
    primary = _get_qualification_line(product_name, spec, province)
    if primary is None:
        primary = record.get("qual_line")
    secondary = _get_secondary_threshold(product_name, province, spec)
    if not _is_1998_product(product_name) and not product_name:
        secondary = record.get("secondary_line", secondary)
    return primary, secondary


def _threshold_header_label(records, index):
    """Format a header label while making mixed-province rules explicit."""
    values = set()
    for record in records:
        threshold = _get_record_thresholds(record)[index]
        if threshold is not None:
            values.add(_format_threshold(threshold))
    if len(values) == 1:
        return f"{next(iter(values))}元"
    return "各省标准"


def _format_threshold(val):
    """格式化阈值用于表头显示：55 -> "55", 74.99 -> "74.99" """
    if val == int(val):
        return str(int(val))
    return f"{val}"


def _get_active_rules(records):
    """从 records 中提取实际出现的产品×规格对应的合格规则（保持 QUALIFICATION_RULES 顺序）

    1998系列（漓泉/铂金/特渠）统一按漓泉1998规则归并。
    """
    active_keys = set()
    for r in records:
        product_name = r.get("product_name", "")
        spec = r.get("spec", "")
        # 1998 系列归一化为"漓泉1998"后匹配
        normalized = product_name
        if _is_1998_product(product_name):
            for kw in ("铂金1998", "特渠1998"):
                normalized = normalized.replace(kw, "漓泉1998")
        for kw_product, kw_spec, line in QUALIFICATION_RULES:
            if kw_product in normalized and kw_spec in spec:
                active_keys.add((kw_product, kw_spec, line))
    # 保持原始顺序
    return [rule for rule in QUALIFICATION_RULES if rule in active_keys]


def _get_province(region):
    """从区域名（如'南宁市'）反查省份"""
    if not region:
        return ""
    # 兼容人工填写的“广东省/广西省”或“广东-广州”等省份前缀。
    for province in _1998_PROVINCE_THRESHOLDS:
        if province in region:
            return province
    if region in CITY_TO_PROVINCE:
        return CITY_TO_PROVINCE[region]
    for city, prov in CITY_TO_PROVINCE.items():
        if city in region:
            return prov
    return ""


def _short_name(full_name):
    """简化产品名：去掉规格部分和'啤酒'后缀"""
    name = re.sub(r'\s*\d+ml.*$', '', full_name)
    name = name.replace("啤酒", "").strip()
    return name


def _extract_images_from_xlsx(xlsx_path):
    """
    从源 xlsx 的 WPS cellimages 中提取所有图片字节

    Returns:
        dict: {img_id: image_bytes}
    """
    images = {}
    with zipfile.ZipFile(xlsx_path, "r") as z:
        names = z.namelist()
        if "xl/cellimages.xml" not in names:
            return images

        ci_xml = z.read("xl/cellimages.xml").decode("utf-8")
        # 解析 ID -> rId 映射
        pairs = re.findall(r'name="(ID_[A-F0-9]+)".*?r:embed="(rId\d+)"', ci_xml)
        id_to_rid = {img_id: rid for img_id, rid in pairs}

        # 解析 rId -> media 路径
        rels_xml = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
        rid_to_media = {}
        for m in re.finditer(r'Id="(rId\d+)".*?Target="(media/[^"]+)"', rels_xml):
            rid_to_media[m.group(1)] = m.group(2)

        # 读取图片字节
        for img_id, rid in id_to_rid.items():
            media_path = rid_to_media.get(rid, "")
            if media_path:
                images[img_id] = z.read(f"xl/{media_path}")

    return images


def _read_data_with_images(xlsx_path):
    """
    从巡查表读取所有数据行 + 对应图片字节

    使用 M 列"产品理论成交价格"（= 成交价 - 打包配送费）作为合格率判定和统计价格。
    同时读取全部 A~P 列原始数据，用于核查明细表原样展示。

    Returns:
        list[dict]: 每条记录包含 province/region/shop_name/product_name/spec/
                    original_price/final_price/theory_price/qual_line/passed/image_bytes
                    以及 all_columns (list, A~P 列原始值)
    """
    images = _extract_images_from_xlsx(xlsx_path)
    # data_only=True 读取数值（公式缓存值），但对于 O 列 DISPIMG 公式，
    # 某些文件缓存值为空(<v></v>)导致读到 None，需要同时加载 data_only=False 版本
    # 以读取公式字符串来提取图片 ID
    wb = load_workbook(xlsx_path, data_only=True)
    wb_formulas = load_workbook(xlsx_path, data_only=False)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_formulas = wb_formulas[sheet_name] if sheet_name in wb_formulas.sheetnames else None
        for row in range(3, ws.max_row + 1):
            region = ws.cell(row=row, column=2).value       # B: 所属区域
            shop_name = ws.cell(row=row, column=4).value     # D: 店铺名称
            product_name = ws.cell(row=row, column=5).value  # E: 产品名称
            original_price = ws.cell(row=row, column=6).value  # F: 原价
            final_price = ws.cell(row=row, column=7).value   # G: 成交价
            theory_price = ws.cell(row=row, column=13).value  # M: 产品理论成交价格

            if not product_name or not shop_name:
                continue
            if final_price is None:
                continue

            spec = _extract_spec(str(product_name))

            try:
                fp = float(final_price)
            except (ValueError, TypeError):
                continue

            try:
                orig = float(original_price) if original_price else 0
            except (ValueError, TypeError):
                orig = 0

            province = _get_province(str(region) if region else "")
            qual_line = _get_qualification_line(str(product_name), spec, province)
            if qual_line is None:
                continue
            secondary_line = _get_secondary_threshold(
                str(product_name), province, spec)

            # 提取图片
            # O 列是 DISPIMG 公式，data_only=True 时可能读到 None（缓存值为空）
            # 需要从 data_only=False 版本读取公式字符串提取图片 ID
            img_bytes = None
            o_value = ws.cell(row=row, column=15).value  # O: 图片 (data_only=True)
            if not o_value and ws_formulas:
                o_value = ws_formulas.cell(row=row, column=15).value  # 公式字符串
            if o_value:
                m = re.search(r"(ID_[A-F0-9]+)", str(o_value))
                if m:
                    img_bytes = images.get(m.group(1))

            # 读取全部 A~P 列原始值（用于核查明细表原样展示）
            # M列(理论成交价)和N列(去除平台优惠价)在源文件中是公式，
            # data_only=True 可能读到缓存值也可能读到None
            # 需要手动计算确保有值
            all_columns = []
            for col in range(1, 17):
                val = ws.cell(row=row, column=col).value
                all_columns.append(val)

            # 手动计算 M列(理论成交价 = G - L) 和 N列(去除平台优惠价 = G + J + K - L)
            # 确保即使源文件公式无缓存值也能正确显示
            g_val = _safe_float(all_columns[6])   # G: 成交价
            j_val = _safe_float(all_columns[9])   # J: 优惠券
            k_val = _safe_float(all_columns[10])  # K: 红包
            l_val = _safe_float(all_columns[11])  # L: 配送费

            # 理论成交价：优先用 M 列缓存值，无缓存值时用 G-L 手动计算
            m_val = all_columns[12]  # M
            if m_val is not None:
                try:
                    tp = float(m_val)
                except (ValueError, TypeError):
                    tp = round(g_val - l_val, 2)
                    m_val = tp
            else:
                tp = round(g_val - l_val, 2)
                m_val = tp
            all_columns[12] = m_val

            n_val = all_columns[13]  # N
            if n_val is None:
                n_val = round(g_val + j_val + k_val - l_val, 2)
            all_columns[13] = n_val

            # O列(图片)不在此存值，由 _build_detail_sheet 用 DISPIMG 占位符处理
            all_columns[14] = None

            # 允许0.1元误差，59.9及以上算合格
            passed = tp >= qual_line - 0.1

            records.append({
                "province": province,
                "region": str(region) if region else "",
                "shop_name": str(shop_name),
                "product_name": str(product_name),
                "spec": spec,
                "original_price": orig,
                "final_price": fp,
                "theory_price": tp,
                "qual_line": qual_line,
                "secondary_line": secondary_line,
                "passed": passed,
                "image_bytes": img_bytes,
                "all_columns": all_columns,
            })

    return records


# =========================================================
# Sheet 构建
# =========================================================

def _build_province_summary(ws, records, start_row=1):
    """
    构建"分省汇总"区块，从 start_row 开始写入

    Returns:
        int: 下一个可用行号
    """
    col_widths = [8, 14, 14, 10, 8, 8, 10, 10, 14, 14, 16, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 根据实际数据确定产品类型和阈值
    active_rules = _get_active_rules(records)
    if not active_rules:
        return start_row
    # 混合省份时表头使用“各省标准”，具体阈值写在每个省份行和标准说明中。
    primary_label = _threshold_header_label(records, 0)
    secondary_label = _threshold_header_label(records, 1)

    # 标题
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=14)
    ws.cell(row=start_row, column=1, value="一、分省价格合格率汇总表")
    _apply_style(ws.cell(row=start_row, column=1), STYLE_TITLE)

    # 合格标准行（只显示实际出现的产品规格，产品名取自实际数据以兼容铂金/特渠命名）
    # 同一规格可能有多个产品名变体（铂金1998/特渠1998/漓泉1998），合并显示
    seen_name_spec = []  # [(province, product_name, spec, primary, secondary)]
    name_spec_set = set()
    for r in records:
        nm = _short_name(r["product_name"])
        sp = r["spec"]
        primary, secondary = _get_record_thresholds(r)
        key = (r.get("province", ""), nm, sp, primary, secondary)
        if key not in name_spec_set:
            name_spec_set.add(key)
            seen_name_spec.append(key)
    std_parts = [
        f"{province or '未识别省份'}{nm}（{sp.replace('500ml*', '')}）≥{primary}元；"
        f"{secondary}元以上计入第二档"
        for province, nm, sp, primary, secondary in seen_name_spec
    ]
    std_text = "合格标准：  " + "  |  ".join(std_parts) + "\n理论成交价总部定义：产品理论成交价格=产品成交价格-打包、配送费"
    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1, end_column=14)
    ws.cell(row=start_row + 1, column=1, value=std_text)
    _apply_style(ws.cell(row=start_row + 1, column=1), STYLE_SUBTITLE)

    # 表头（动态显示合格线和第二档阈值）
    header_row = start_row + 2
    headers = ["省份", "产品名称", "规格", "合格线\n(元)", "总数",
               f"合格数\n（{primary_label}以上）",
               "不合格数", f"合格率\n（{primary_label}以上售价）",
               f"{secondary_label}以上\n价格", f"{secondary_label}以下\n价格",
               f"合格率\n（{secondary_label}以上售价）",
               "最低理论\n成交价", "最高理论\n成交价", "平均理论\n成交价"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        _apply_style(cell, STYLE_HEADER)

    # 按省份×产品×规格分组
    groups = defaultdict(list)
    for r in records:
        key = (r["province"], r["product_name"], r["spec"])
        groups[key].append(r)

    row_idx = header_row + 1
    total_count = 0
    total_pass = 0
    total_fail = 0
    total_above = 0
    total_below = 0
    first_data_row = row_idx
    last_data_row = row_idx  # 至少有一行数据时更新

    for (province, full_name, spec), items in sorted(groups.items()):
        count = len(items)
        primary_line, secondary_threshold = _get_record_thresholds(items[0])
        passed = sum(
            1 for r in items
            if r["theory_price"] >= _get_record_thresholds(r)[0] - 0.1
        )
        failed = count - passed
        prices = [r["theory_price"] for r in items]
        avg_price = round(sum(prices) / len(prices), 1) if prices else 0

        # 第二档阈值按省份执行（1998 广东65/广西55；U8为55）。
        above_items = [
            r for r in items
            if r["theory_price"] >= _get_record_thresholds(r)[1]
        ]
        below_items = [r for r in items if r["theory_price"] < secondary_threshold]
        above_count = len(above_items)
        below_count = len(below_items)

        # 合格率/第二档合格率改用 Excel 公式：F/E、I/E
        # （第二档合格率 = 第二档以上数量 / 总数，与用户修改的表一致）
        values = [province, _short_name(full_name), spec, primary_line,
                  count, passed, failed,
                  f"=F{row_idx}/E{row_idx}",
                  above_count, below_count,
                  f"=I{row_idx}/E{row_idx}",
                  min(prices), max(prices), avg_price]

        rate = passed / count if count > 0 else 0
        above_rate = above_count / count if count > 0 else 0
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 8:  # 合格率
                _apply_style(cell, STYLE_PASS if rate >= 1.0 else STYLE_FAIL)
            elif col == 11:  # 第二档合格率
                _apply_style(cell, STYLE_PASS if above_rate >= 1.0 else STYLE_FAIL)
            else:
                _apply_style(cell, STYLE_DATA)
        ws.cell(row=row_idx, column=8).number_format = "0%"
        ws.cell(row=row_idx, column=11).number_format = "0%"

        total_count += count
        total_pass += passed
        total_fail += failed
        total_above += above_count
        total_below += below_count
        last_data_row = row_idx
        row_idx += 1

    # 总计行（用公式引用上方数据区域：F/E、I/E）
    total_row = row_idx
    total_rate = total_pass / total_count if total_count > 0 else 0
    total_above_rate = total_above / total_count if total_count > 0 else 0
    totals = ["总计", "", "", "", total_count, total_pass, total_fail,
              f"=F{total_row}/E{total_row}",
              total_above, total_below,
              f"=I{total_row}/E{total_row}",
              "", "", ""]
    for col, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        if col == 8:
            _apply_style(cell, STYLE_PASS if total_rate >= 1.0 else STYLE_FAIL)
        elif col == 11:
            _apply_style(cell, STYLE_PASS if total_above_rate >= 1.0 else STYLE_FAIL)
        else:
            _apply_style(cell, STYLE_TOTAL)
    ws.cell(row=total_row, column=8).number_format = "0%"
    ws.cell(row=total_row, column=11).number_format = "0%"

    # 行高
    ws.row_dimensions[start_row].height = 30
    ws.row_dimensions[start_row + 1].height = 34
    ws.row_dimensions[header_row].height = 42
    for r in range(header_row + 1, row_idx + 1):
        ws.row_dimensions[r].height = 22

    return row_idx + 1  # 留1行空隙


def _build_city_summary(ws, records, start_row=1):
    """
    构建"分地级市汇总"区块，从 start_row 开始写入

    Returns:
        int: 下一个可用行号
    """
    col_widths = [8, 10, 14, 14, 10, 8, 8, 10, 10, 14, 14, 16, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 根据实际数据确定产品类型和阈值
    active_rules = _get_active_rules(records)
    if not active_rules:
        return start_row
    primary_label = _threshold_header_label(records, 0)
    secondary_label = _threshold_header_label(records, 1)

    # 标题
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=15)
    ws.cell(row=start_row, column=1, value="二、分地级市价格合格率汇总表")
    _apply_style(ws.cell(row=start_row, column=1), STYLE_TITLE)

    # 合格标准行（只显示实际出现的产品规格，产品名取自实际数据以兼容铂金/特渠命名）
    seen_name_spec = []
    name_spec_set = set()
    for r in records:
        nm = _short_name(r["product_name"])
        sp = r["spec"]
        primary, secondary = _get_record_thresholds(r)
        key = (r.get("province", ""), nm, sp, primary, secondary)
        if key not in name_spec_set:
            name_spec_set.add(key)
            seen_name_spec.append(key)
    std_parts = [
        f"{province or '未识别省份'}{nm}（{sp.replace('500ml*', '')}）≥{primary}元；"
        f"{secondary}元以上计入第二档"
        for province, nm, sp, primary, secondary in seen_name_spec
    ]
    std_text = "合格标准：" + "  |  ".join(std_parts) + "\n理论成交价总部定义：产品理论成交价格=产品成交价格-打包、配送费"
    ws.merge_cells(start_row=start_row + 1, start_column=1,
                   end_row=start_row + 1, end_column=15)
    ws.cell(row=start_row + 1, column=1, value=std_text)
    _apply_style(ws.cell(row=start_row + 1, column=1), STYLE_SUBTITLE)

    # 表头（动态显示合格线和第二档阈值）
    header_row = start_row + 2
    headers = ["省份", "地级市", "产品名称", "规格", "合格线\n(元)", "总数",
               f"合格数\n（{primary_label}以上）",
               "不合格数", f"合格率\n（{primary_label}以上售价）",
               f"{secondary_label}以上\n价格", f"{secondary_label}以下\n价格",
               f"合格率\n（{secondary_label}以上售价）",
               "最低理论\n成交价", "最高理论\n成交价", "平均理论\n成交价"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        _apply_style(cell, STYLE_HEADER)

    # 按省份×城市×产品×规格分组
    groups = defaultdict(list)
    for r in records:
        key = (r["province"], r["region"], r["product_name"], r["spec"])
        groups[key].append(r)

    row_idx = header_row + 1
    for (province, region, full_name, spec), items in sorted(groups.items()):
        count = len(items)
        primary_line, secondary_threshold = _get_record_thresholds(items[0])
        passed = sum(
            1 for r in items
            if r["theory_price"] >= _get_record_thresholds(r)[0] - 0.1
        )
        failed = count - passed
        prices = [r["theory_price"] for r in items]
        avg_price = round(sum(prices) / len(prices), 1) if prices else 0

        # 第二档阈值按省份执行（1998 广东65/广西55；U8为55）。
        above_items = [
            r for r in items
            if r["theory_price"] >= _get_record_thresholds(r)[1]
        ]
        below_items = [r for r in items if r["theory_price"] < secondary_threshold]
        above_count = len(above_items)
        below_count = len(below_items)

        # 合格率/第二档合格率改用 Excel 公式：G/F、J/F
        # （第二档合格率 = 第二档以上数量 / 总数，与用户修改的表一致）
        values = [province, region, _short_name(full_name), spec, primary_line,
                  count, passed, failed,
                  f"=G{row_idx}/F{row_idx}",
                  above_count, below_count,
                  f"=J{row_idx}/F{row_idx}",
                  min(prices), max(prices), avg_price]

        rate = passed / count if count > 0 else 0
        above_rate = above_count / count if count > 0 else 0
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 9:  # 合格率
                _apply_style(cell, STYLE_PASS if rate >= 1.0 else STYLE_FAIL)
            elif col == 12:  # 第二档合格率
                _apply_style(cell, STYLE_PASS if above_rate >= 1.0 else STYLE_FAIL)
            else:
                _apply_style(cell, STYLE_DATA)
        ws.cell(row=row_idx, column=9).number_format = "0%"
        ws.cell(row=row_idx, column=12).number_format = "0%"
        row_idx += 1

    # 行高
    ws.row_dimensions[start_row].height = 30
    ws.row_dimensions[start_row + 1].height = 34
    ws.row_dimensions[header_row].height = 42
    for r in range(header_row + 1, row_idx):
        ws.row_dimensions[r].height = 22

    return row_idx + 1  # 留1行空隙


def _build_detail_sheet(ws, records, start_row=1):
    """
    构建"核查明细表"（完整巡查表格式 A~P 列），不合格行整行标红

    Returns:
        tuple: (next_row, image_records)
    """
    # 列宽（与模板一致）
    col_widths = {
        "A": 33, "B": 11, "C": 12, "D": 25, "E": 19,
    }
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    # 标题行
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=16)
    title_cell = ws.cell(row=start_row, column=1,
                         value="燕京啤酒全国即时零售渠道产品价格巡查表")
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True, color=Color(theme=0))
    title_cell.fill = PatternFill(fill_type="solid", fgColor="FF003366")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 38

    # 表头行
    header_row = start_row + 1
    headers = [
        "分公司", "所属主要区域", "区域内即时零售平台", "店铺名称", "平台在售燕京产品",
        "产品原价\n（商品标价）", "产品成交单价\n（最终付款价格）",
        "商品优惠/商品活动\n（店铺活动）", "满减活动\n（店铺活动）",
        "优惠卷\n（平台下发）", "红包\n（平台下发）", "打包、配送费",
        "产品理论成交价格\n（产品成交价格-打包、配送费）",
        "去除平台优惠价格\n（最终付款价格-打包、配送费+优惠卷+红包）",
        "图片", "备注",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=Color(theme=0))
        cell.fill = PatternFill(fill_type="solid", fgColor="FF0070C0")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = copy(_THIN_BORDER)
    ws.row_dimensions[header_row].height = 38

    # 数据行
    row_idx = header_row + 1
    image_records = []

    # 按省份->城市->店铺排序
    sorted_records = sorted(records, key=lambda r: (r["province"], r["region"], r["shop_name"]))

    for r in sorted_records:
        # 写入 A~P 列（共16列）。M/N 的公式会在写入后按当前输出行重建，
        # 不能复用来源行号，否则排序后会引用错误的数据行。
        cols = r["all_columns"]  # list of 16 values
        for col_idx in range(1, 17):
            val = cols[col_idx - 1] if col_idx - 1 < len(cols) else None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            # 基础样式
            cell.font = Font(name=FONT_NAME, size=10, bold=False)
            cell.fill = PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = copy(_THIN_BORDER)

            # 不合格行：整行字体标红
            if not r["passed"]:
                cell.font = Font(name=FONT_NAME, size=10, bold=False, color="FFFF0000")

        ws.cell(row=row_idx, column=13).value = f"=G{row_idx}-L{row_idx}"
        ws.cell(row=row_idx, column=14).value = f"=G{row_idx}+J{row_idx}+K{row_idx}-L{row_idx}"

        # O 列（图片）：用 DISPIMG 占位符
        o_cell = ws.cell(row=row_idx, column=15)
        if r["image_bytes"]:
            placeholder = f"PH_{uuid.uuid4().hex.upper()}"
            o_cell.value = f'=_xlfn.DISPIMG("{placeholder}",1)'
            image_records.append((row_idx, r["image_bytes"], placeholder))

        ws.row_dimensions[row_idx].height = 38
        row_idx += 1

    return row_idx, image_records


# =========================================================
# WPS cellimages 注入
# =========================================================

def _inject_wps_cellimages(xlsx_path, image_records):
    """
    向 xlsx 文件注入 WPS 格式的嵌入图片（cellimages）
    图片以字节形式提供，嵌入单元格内。

    Args:
        xlsx_path: xlsx 文件路径
        image_records: list of (row_number, image_bytes, placeholder)
    """
    if not image_records:
        return

    cellimage_xmls = []
    rel_entries = []       # (rId, target)
    media_entries = []     # (media_filename, image_bytes)
    placeholder_map = {}   # placeholder -> img_id

    for i, (row, img_bytes, placeholder) in enumerate(image_records):
        img_id = f"ID_{uuid.uuid4().hex.upper()}"
        placeholder_map[placeholder] = img_id
        rid = f"rId{i + 1}"
        media_filename = f"image_sum_{i + 1}.jpeg"

        # 获取图片尺寸（EMU 单位，1 px = 9525 EMU）
        try:
            pil = PILImage.open(io.BytesIO(img_bytes))
            w_px, h_px = pil.size
        except Exception:
            w_px, h_px = 200, 400

        cx = w_px * 9525
        cy = h_px * 9525

        cellimage_xmls.append(
            f'<etc:cellImage><xdr:pic>'
            f'<xdr:nvPicPr><xdr:cNvPr id="{i + 2}" name="{img_id}"/>'
            f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
            f'</xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="{rid}"/>'
            f'<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
            f'</xdr:pic></etc:cellImage>'
        )

        rel_entries.append((rid, f"media/{media_filename}"))
        media_entries.append((media_filename, img_bytes))

    # 构建 cellimages.xml
    cellimages_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<etc:cellImages '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">'
        + "".join(cellimage_xmls) +
        '</etc:cellImages>'
    )

    # 构建 cellimages.xml.rels
    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="{target}"/>'
            for rid, target in rel_entries
        )
        + '</Relationships>'
    )

    # 重新打包 xlsx
    tmp_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                name = item.filename

                # 跳过已有的 cellimages
                if name == "xl/cellimages.xml":
                    continue
                if name == "xl/_rels/cellimages.xml.rels":
                    continue

                data = zin.read(name)

                # 修改 [Content_Types].xml
                if name == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    if "cellimage" not in text:
                        text = text.replace(
                            "</Types>",
                            '<Override PartName="/xl/cellimages.xml" '
                            'ContentType="application/vnd.wps-officedocument.cellimage+xml"/>'
                            "</Types>",
                        )
                    data = text.encode("utf-8")

                # 修改 workbook.xml.rels
                if name == "xl/_rels/workbook.xml.rels":
                    text = data.decode("utf-8")
                    if "cellimages" not in text:
                        text = text.replace(
                            "</Relationships>",
                            f'<Relationship Id="rId_cellimages" '
                            f'Type="http://www.wps.cn/officeDocument/2020/cellImage" '
                            f'Target="cellimages.xml"/>'
                            f"</Relationships>",
                        )
                    data = text.encode("utf-8")

                # 替换 sharedStrings.xml 中的占位符
                if name == "xl/sharedStrings.xml":
                    text = data.decode("utf-8")
                    for placeholder, img_id in sorted(placeholder_map.items(), key=lambda x: -len(x[0])):
                        text = text.replace(placeholder, img_id)
                    data = text.encode("utf-8")

                # 替换 worksheets 中的占位符
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    text = data.decode("utf-8")
                    for placeholder, img_id in sorted(placeholder_map.items(), key=lambda x: -len(x[0])):
                        text = text.replace(placeholder, img_id)
                    data = text.encode("utf-8")

                zout.writestr(item, data)

            # 写入新文件（图片已压缩，用 STORED 避免无收益的二次 deflate）
            zout.writestr("xl/cellimages.xml", cellimages_xml.encode("utf-8"))
            zout.writestr("xl/_rels/cellimages.xml.rels", rels_xml.encode("utf-8"))
            for filename, img_bytes in media_entries:
                zi = zipfile.ZipInfo(f"xl/media/{filename}")
                zi.compress_type = zipfile.ZIP_STORED
                zout.writestr(zi, img_bytes)

    shutil.move(tmp_path, xlsx_path)


# =========================================================
# 主函数
# =========================================================

def generate_summary_report(xlsx_path, output_dir=None, provinces=None, product_type=None):
    """
    从巡查表生成分省/分地级市汇总 + 明细表(含截图)

    Args:
        xlsx_path: 巡查表 xlsx 路径
        output_dir: 输出目录，默认为巡查表同目录
        provinces: 选中的省份列表，None 表示包含所有省份
        product_type: 产品类型筛选，"u8" 只统计燕京U8，"1998" 只统计漓泉1998，
                      None 表示不筛选（按实际数据自动判断）

    Returns:
        str: 生成的汇总表文件路径
    """
    # 1. 读取数据 + 图片
    records = _read_data_with_images(xlsx_path)
    if not records:
        raise ValueError("未从巡查表中读取到有效数据，请检查文件内容")

    # 2. 按省份筛选
    if provinces is not None and len(provinces) > 0:
        records = [r for r in records if r["province"] in provinces]
        if not records:
            raise ValueError(f"选中的省份 {provinces} 下没有匹配的记录")

    # 3. 按产品类型筛选
    if product_type == "u8":
        records = [r for r in records if "U8" in r["product_name"] or "燕京" in r["product_name"]]
        if not records:
            raise ValueError("巡查表中未找到燕京U8产品数据")
    elif product_type == "1998":
        records = [r for r in records if _is_1998_product(r["product_name"])]
        if not records:
            raise ValueError("巡查表中未找到漓泉1998产品数据")

    # 4. 创建工作簿
    wb = Workbook()

    # Sheet1: 汇总表（分省 + 分地级市，上下排列）
    ws_summary = wb.active
    ws_summary.title = "汇总表"

    # 区块1: 分省汇总
    next_row = _build_province_summary(ws_summary, records, start_row=1)

    # 区块2: 分地级市汇总
    next_row = _build_city_summary(ws_summary, records, start_row=next_row)

    # Sheet2: 核查明细表（完整巡查表格式，不合格标红）
    ws_detail = wb.create_sheet("核查明细表")
    _, image_records = _build_detail_sheet(ws_detail, records, start_row=1)

    # 4. 确定输出路径
    if output_dir is None:
        output_dir = os.path.dirname(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"价格合格率汇总表_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    # 5. 保存
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output_path)

    # 6. 注入 WPS 图片
    if image_records:
        _inject_wps_cellimages(output_path, image_records)

    return output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法: python summary_generator.py <巡查表.xlsx> [省份1,省份2,...]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    provinces = sys.argv[2].split(",") if len(sys.argv) > 2 else None

    try:
        result = generate_summary_report(xlsx_path, provinces=provinces)
        print(f"已生成: {result}")
    except Exception as e:
        print(f"错误: {e}")
