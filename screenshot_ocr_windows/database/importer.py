"""
店铺/城市智能匹配数据库 — Importer（Excel 完整表格投喂）

流程：读取 -> 清洗 -> 匹配已有店铺 -> 更新别名/城市 -> 记录学习来源 -> 生成投喂报告

幂等：同一文件（SHA-256）重复投喂直接识别为已有批次，不重复插入。
冲突：同店不同城市 -> city_matches 标记 CONFLICT，不做静默覆盖。
"""

import hashlib
import logging
import os
import re

from openpyxl import load_workbook

from . import repository
from .learner import learn_correction

# Batch rows are created before corrections so every correction has an audit parent.
from .matcher import normalize

logger = logging.getLogger(__name__)

# 投喂 Excel 的列（与巡查表 A~P 对应，0-based）
COL_REGION = 1    # B: 所属区域（城市）
COL_SHOP = 3      # D: 店铺名称

_REGION_HEADERS = {"区域", "所属区域", "所属主要区域"}
_SHOP_HEADERS = {"店铺", "店铺名称"}
_NON_SHOP_NAMES = {
    "规格", "产品规格", "合格线", "合格线元", "原价", "成交价", "产品原价",
    "产品成交单价", "商品标价", "备注", "产品", "店名", "城市", "所属区域",
}
_PRODUCT_SPEC_PATTERN = re.compile(
    r"(?:\d+(?:\.\d+)?\s*(?:ml|毫升|l|升|g|克|kg|千克|斤)|"
    r"\d+\s*[x*×]\s*\d+\s*(?:听|瓶|罐|箱|包|袋|支|个))",
    re.IGNORECASE,
)


def _file_hash(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _extract_city(region_text):
    """从区域文本提取城市名（'三亚市' / '三亚' -> 三亚市；空 -> ''）"""
    if not region_text:
        return ""
    s = str(region_text).strip()
    if not s:
        return ""
    if not s.endswith("市"):
        s = s + "市"
    return s


def _header_text(value):
    """Normalize a header to tolerate whitespace and parenthesis variants."""
    return re.sub(r"[\s（）()]", "", str(value or "")).strip()


def _is_inspection_sheet(sheet):
    """Accept only detail sheets that expose region and shop columns."""
    for row in (2, 1):
        region_header = _header_text(
            sheet.cell(row=row, column=COL_REGION + 1).value)
        shop_header = _header_text(
            sheet.cell(row=row, column=COL_SHOP + 1).value)
        if region_header in _REGION_HEADERS and shop_header in _SHOP_HEADERS:
            return True
    return False


def is_invalid_shop_name(value):
    """Reject headers, standalone numeric values and product specifications."""
    name = str(value or "").strip()
    normalized = _header_text(name).lower()
    if not normalized or normalized in _NON_SHOP_NAMES:
        return True
    if re.fullmatch(r"\d+(?:\.\d+)?", normalized):
        return True
    return bool(_PRODUCT_SPEC_PATTERN.search(normalized))


def import_excel(excel_path, operator="gui"):
    """投喂一份人工修正后的 Excel 巡查表。

    Returns:
        dict: 投喂报告
          {batch_id, status, filename, total_rows, new_shops, new_aliases,
           updated_shops, updated_cities, conflicts, ignored_rows,
           duplicate: bool, detail: {...}}
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel 不存在: {excel_path}")

    fhash = _file_hash(excel_path)

    # 幂等：同一文件已投喂过 -> 直接返回历史批次
    existing = repository.get_batch_by_hash(fhash)
    if existing:
        return {
            "batch_id": existing["batch_id"],
            "status": "duplicate",
            "filename": existing["filename"],
            "duplicate": True,
            "import_time": existing["import_time"],
            "total_rows": existing["total_rows"],
            "new_shops": existing["new_shops"],
            "new_aliases": existing["new_aliases"],
            "updated_shops": existing["updated_shops"],
            "updated_cities": existing["updated_cities"],
            "conflicts": existing["conflicts"],
            "ignored_rows": existing["ignored_rows"],
        }

    wb = load_workbook(excel_path, data_only=True)
    batch_id = repository.add_batch(filename=os.path.basename(excel_path), file_hash=fhash, operator=operator, status="processing")

    report = {
        "total_rows": 0, "new_shops": 0, "new_aliases": 0,
        "updated_shops": 0, "updated_cities": 0, "conflicts": 0,
        "ignored_rows": 0,
        "detail": {
            "created": [], "updated": [], "conflict": [], "ignored": [],
            "skipped_sheets": [],
        },
    }

    for sheet in wb.worksheets:
        if not _is_inspection_sheet(sheet):
            report["detail"]["skipped_sheets"].append(sheet.title)
            logger.info("skip non-inspection worksheet: %s", sheet.title)
            continue
        for row in range(3, sheet.max_row + 1):  # 第3行开始是数据
            shop_val = sheet.cell(row=row, column=COL_SHOP + 1).value
            region_val = sheet.cell(row=row, column=COL_REGION + 1).value
            if not shop_val:
                continue
            shop_name = str(shop_val).strip()
            if not shop_name:
                continue

            city = _extract_city(region_val)
            # 清洗后过短视为无效行
            if len(normalize(shop_name)) < 2 or is_invalid_shop_name(shop_name):
                report["ignored_rows"] += 1
                report["detail"]["ignored"].append(shop_name)
                continue

            report["total_rows"] += 1
            try:
                out = learn_correction(
                    shop_name, shop_name, city, operator=operator,
                    source="import", batch_id=batch_id,
                )
                if out["created"]:
                    report["new_shops"] += 1
                    report["new_aliases"] += 1
                    report["detail"]["created"].append(out["canonical_name"])
                else:
                    report["updated_shops"] += 1
                    report["detail"]["updated"].append(out["canonical_name"])
                if out["conflict"]:
                    report["conflicts"] += 1
                    report["updated_cities"] += 1
                    report["detail"]["conflict"].append(out["canonical_name"])
                elif city:
                    report["updated_cities"] += 1
            except Exception as e:
                logger.warning("import row error %s: %s", shop_name, e)
                report["ignored_rows"] += 1

    repository.update_batch(batch_id, total_rows=report["total_rows"], new_shops=report["new_shops"], new_aliases=report["new_aliases"], updated_shops=report["updated_shops"], updated_cities=report["updated_cities"], conflicts=report["conflicts"], ignored_rows=report["ignored_rows"], status="partial_failed" if report["ignored_rows"] else "ok")
    report["batch_id"] = batch_id
    report["status"] = "partial_failed" if report["ignored_rows"] else "ok"
    report["filename"] = os.path.basename(excel_path)
    report["duplicate"] = False
    return report
