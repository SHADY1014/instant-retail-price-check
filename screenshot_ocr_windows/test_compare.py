"""
测试对比脚本

1. 对 /tmp/test_screenshots/ 中的 74 张截图运行 OCR + 解析
2. 自动按品牌分类，写入新 Excel
3. 读取手工参考表，按店铺名匹配，逐字段对比
4. 输出差异报告

用法: python3 test_compare.py
"""

import os
import sys
import glob
from collections import defaultdict

# 确保能 import 同目录模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from ocr_engine import run_ocr
from field_parser import parse_ocr_to_fields, detect_brand
import excel_writer

# 参考表路径（手工填写的）
REF_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "燕京啤酒即时零售渠道价格巡查表 0708.xlsx",
)

# 测试截图目录
SCREENSHOT_DIR = "/tmp/test_screenshots"

# 4个 Sheet 名称
SHEET_NAMES = [
    "1.燕京即时零售渠道价格巡查表",
    "2.雪花即时零售渠道价格巡查表",
    "3.青岛即时零售渠道价格巡查表",
    "4.百威即时零售渠道价格巡查表",
]

BRAND_NAMES = {0: "燕京", 1: "雪花", 2: "青岛", 3: "百威"}


def run_ocr_all(images, default_region="南宁市"):
    """对所有图片运行 OCR + 解析，返回 records 列表"""
    records = []
    total = len(images)

    for i, img in enumerate(images):
        fname = os.path.basename(img)
        print(f"  [{i+1}/{total}] {fname}")

        try:
            ocr_data = run_ocr(img)
            fields = parse_ocr_to_fields(ocr_data, default_region)
            d = fields.to_dict()
            d["image_path"] = img
            records.append(d)
        except Exception as e:
            print(f"    !! OCR 失败: {e}")
            d = {
                "branch_company": "燕京啤酒（桂林漓泉）股份有限公司",
                "region": default_region,
                "platform": "美团闪购",
                "shop_name": "",
                "product_name": "",
                "original_price": 0.0,
                "final_price": 0.0,
                "shop_discount": 0.0,
                "full_reduction": 0.0,
                "coupon": 0.0,
                "red_packet": 0.0,
                "delivery_fee": 0.0,
                "remark": f"OCR失败: {e}",
                "image_path": img,
            }
            records.append(d)

    return records


def read_reference_table(ref_path):
    """
    读取手工参考表，返回 {sheet_idx: [record_dict, ...]}
    """
    from openpyxl import load_workbook

    wb = load_workbook(ref_path, data_only=True)
    ref_data = {}

    for sheet_idx, sheet_name in enumerate(SHEET_NAMES):
        if sheet_name not in wb.sheetnames:
            ref_data[sheet_idx] = []
            continue

        ws = wb[sheet_name]
        rows = []
        # 从第3行开始读取数据
        for row in range(3, ws.max_row + 1):
            shop_name = ws.cell(row=row, column=4).value  # D: 店铺名称
            product_name = ws.cell(row=row, column=5).value  # E: 产品名称
            if not shop_name or not product_name:
                continue

            record = {
                "branch_company": ws.cell(row=row, column=1).value or "",
                "region": ws.cell(row=row, column=2).value or "",
                "platform": ws.cell(row=row, column=3).value or "",
                "shop_name": str(shop_name).strip(),
                "product_name": str(product_name).strip(),
                "original_price": float(ws.cell(row=row, column=6).value or 0),
                "final_price": float(ws.cell(row=row, column=7).value or 0),
                "shop_discount": float(ws.cell(row=row, column=8).value or 0),
                "full_reduction": float(ws.cell(row=row, column=9).value or 0),
                "coupon": float(ws.cell(row=row, column=10).value or 0),
                "red_packet": float(ws.cell(row=row, column=11).value or 0),
                "delivery_fee": float(ws.cell(row=row, column=12).value or 0),
            }
            rows.append(record)

        ref_data[sheet_idx] = rows

    return ref_data


def compare_records(ocr_records, ref_data):
    """
    对比 OCR 识别结果和手工参考表

    匹配策略：按品牌分表 + 店铺名匹配
    """
    # 按 sheet_idx 分组 OCR 结果
    ocr_by_sheet = defaultdict(list)
    for r in ocr_records:
        sheet_idx = detect_brand(r["product_name"])
        ocr_by_sheet[sheet_idx].append(r)

    all_diffs = []

    for sheet_idx in range(4):
        brand = BRAND_NAMES[sheet_idx]
        ref_rows = ref_data.get(sheet_idx, [])
        ocr_rows = ocr_by_sheet.get(sheet_idx, [])

        print(f"\n{'='*60}")
        print(f"品牌: {brand} (Sheet {sheet_idx+1})")
        print(f"  参考表: {len(ref_rows)} 条")
        print(f"  OCR识别: {len(ocr_rows)} 条")

        # 按 shop_name 匹配
        ref_map = {}
        for r in ref_rows:
            # 归一化店铺名（去掉空格、统一括号）
            key = normalize_shop_name(r["shop_name"])
            ref_map[key] = r

        ocr_matched = set()
        for ocr_r in ocr_rows:
            key = normalize_shop_name(ocr_r["shop_name"])
            if key in ref_map:
                ocr_matched.add(key)
                ref_r = ref_map[key]
                diffs = compare_fields(ref_r, ocr_r)
                if diffs:
                    all_diffs.append((brand, ocr_r["shop_name"], diffs))
            else:
                # OCR 识别到但参考表没有
                all_diffs.append((brand, ocr_r["shop_name"], [("店铺", "参考表无此店铺", "OCR有此店铺")]))

        # 参考表有但 OCR 没有匹配的
        for key, ref_r in ref_map.items():
            if key not in ocr_matched:
                # 只报告参考表有但OCR完全没有的情况
                pass  # 有些店铺照片没有是正常的，不需要报告

    return all_diffs


def normalize_shop_name(name):
    """归一化店铺名，便于匹配"""
    if not name:
        return ""
    # 去空格、统一括号
    name = name.strip()
    name = name.replace("（", "(").replace("）", ")")
    name = name.replace(" ", "")
    return name


def compare_fields(ref, ocr):
    """对比两个记录的字段差异"""
    fields_to_compare = [
        ("product_name", "产品名", str),
        ("original_price", "原价", float),
        ("final_price", "成交价", float),
        ("shop_discount", "优惠", float),
        ("full_reduction", "满减", float),
        ("coupon", "券", float),
        ("red_packet", "红包", float),
        ("delivery_fee", "配送费", float),
    ]

    diffs = []
    for key, label, typ in fields_to_compare:
        ref_val = ref.get(key, 0)
        ocr_val = ocr.get(key, 0)

        if typ == float:
            ref_val = float(ref_val or 0)
            ocr_val = float(ocr_val or 0)
            # 允许小误差（0.5元）
            if abs(ref_val - ocr_val) > 0.5:
                diffs.append((label, ref_val, ocr_val))
        else:
            ref_val = str(ref_val or "").strip()
            ocr_val = str(ocr_val or "").strip()
            if ref_val != ocr_val:
                diffs.append((label, ref_val, ocr_val))

    return diffs


def main():
    print("=" * 60)
    print("美团截图 OCR 测试对比工具")
    print("=" * 60)

    # 1. 检查截图目录
    images = sorted(glob.glob(os.path.join(SCREENSHOT_DIR, "*.jpg")))
    if not images:
        # 也尝试 png
        images = sorted(glob.glob(os.path.join(SCREENSHOT_DIR, "*.png")))
    if not images:
        print(f"错误: 在 {SCREENSHOT_DIR} 中未找到截图文件")
        sys.exit(1)

    print(f"\n找到 {len(images)} 张截图")

    # 2. 运行 OCR
    print(f"\n开始 OCR 识别...")
    records = run_ocr_all(images)

    # 3. 写入 Excel
    print(f"\n写入 Excel（自动分表）...")
    output_path = excel_writer.write_all_brands(records, output_name="测试对比_OCR结果.xlsx")
    print(f"已保存到: {output_path}")

    # 4. 读取参考表
    print(f"\n读取手工参考表: {REF_PATH}")
    if not os.path.exists(REF_PATH):
        print(f"错误: 参考表不存在")
        sys.exit(1)
    ref_data = read_reference_table(REF_PATH)

    for sheet_idx in range(4):
        brand = BRAND_NAMES[sheet_idx]
        count = len(ref_data.get(sheet_idx, []))
        print(f"  {brand}: {count} 条参考数据")

    # 5. 对比
    print(f"\n开始对比...")
    all_diffs = compare_records(records, ref_data)

    # 6. 输出差异报告
    print(f"\n{'='*60}")
    print(f"差异报告")
    print(f"{'='*60}")

    if not all_diffs:
        print("✅ 所有匹配数据完全一致！")
    else:
        print(f"共发现 {len(all_diffs)} 处差异:\n")
        for brand, shop, diffs in all_diffs:
            print(f"  [{brand}] {shop}")
            for field, ref_val, ocr_val in diffs:
                print(f"    {field}: 参考={ref_val} | OCR={ocr_val}")
            print()

    # 统计
    print(f"{'='*60}")
    print(f"统计:")
    print(f"  OCR 总识别: {len(records)} 张")
    brand_counts = defaultdict(int)
    for r in records:
        brand_counts[BRAND_NAMES[detect_brand(r["product_name"])]] += 1
    for brand, count in sorted(brand_counts.items()):
        print(f"    {brand}: {count} 张")
    print(f"  差异数量: {len(all_diffs)}")


if __name__ == "__main__":
    main()
