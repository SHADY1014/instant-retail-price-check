"""巡查表转换为总部供货渠道门店价格检查表。

这个模块刻意不依赖 Qt。解析、店名归一化、合并、判定和输出均通过
小而稳定的函数完成，macOS 和 Windows 桌面端可以直接共用同一套规则。
图片采用 WPS 的 ``DISPIMG``/``cellimages.xml`` 机制，避免在 WPS 中显示
``#REF!``。
"""

from __future__ import annotations

import os
import re
import shutil
import zipfile
from collections import OrderedDict
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import PurePosixPath
from typing import Any, Iterable, Mapping, Optional
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


CONVERTER_VERSION = "1.0.0"
PLATFORMS = ("美团闪购", "淘宝闪购", "京东秒送")
SOURCE_PLATFORM_ALIASES = {
    "美团": "美团闪购",
    "美团闪购": "美团闪购",
    "淘宝": "淘宝闪购",
    "淘宝闪购": "淘宝闪购",
    "京东闪送": "京东秒送",
    "京东秒送": "京东秒送",
    "京东": "京东秒送",
}
SPEC_ORDER = {
    "500ml*6听": 0,
    "500ml*6瓶": 1,
    "500ml*12听": 2,
    "500ml*12瓶": 3,
}
DEFAULT_NO_SALE_TEXT = "无售整件 6听/瓶售卖"

# U8 使用巡查表转总部表时的默认合格线；1998 的省份口径单独配置。
DEFAULT_THRESHOLDS = {
    "u8": {6: 29.9, 12: 59.9},
    "1998": {6: 29.9, 12: 59.9},
}
DEFAULT_1998_PROVINCE_THRESHOLDS = {
    # 1998 的 12 瓶/听采用省份口径；65/55 是汇总表第二档线，
    # 不用于“是否违规”的一票否决判定。
    "广东": {12: 70.0},
    "广西": {12: 60.0},
}

CITY_TO_PROVINCE = {
    city: province
    for province, cities in {
        "广东": "广州 深圳 珠海 汕头 佛山 韶关 湛江 肇庆 江门 茂名 惠州 梅州 汕尾 河源 阳江 清远 东莞 中山 潮州 揭阳 云浮".split(),
        "广西": "南宁 柳州 桂林 梧州 北海 防城港 钦州 贵港 玉林 百色 贺州 河池 来宾 崇左".split(),
        "海南": "海口 三亚 三沙 儋州".split(),
        "贵州": "贵阳 六盘水 遵义 安顺 毕节 铜仁".split(),
        "云南": "昆明 曲靖 玉溪 保山 昭通 丽江 普洱 临沧".split(),
    }.items()
    for city_name in cities
    for city in (city_name, city_name + "市")
}


@dataclass
class SourceRecord:
    """源表一行的标准化表示。"""

    row_number: int
    region: str
    shop_name: str
    platform: str
    product_name: str
    spec: str
    final_price: Optional[float]
    theory_price: Optional[float]
    image_bytes: Optional[bytes] = None
    shop_key: str = ""
    display_name: str = ""
    is_no_sale: bool = False


@dataclass
class StoreCheckRow:
    """输出表的一行。"""

    shop_name: str
    product_name: str
    status: str
    platform_values: dict[str, Optional[SourceRecord]] = field(
        default_factory=dict
    )
    no_sale: bool = False
    shop_key: str = ""
    spec: str = ""
    threshold: Optional[float] = None


@dataclass
class ConversionResult:
    """转换结果及审计信息。"""

    output_path: str
    rows: list[StoreCheckRow]
    pending: list[str]
    log_entries: list[str]


def _safe_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"/", "-", "—"}:
        return None
    text = re.sub(r"[¥￥$元块,，\s]", "", text)
    text = text.translate(str.maketrans("０１２３４５６７８９．", "0123456789."))
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _clean_shop_name(shop_name: str) -> str:
    """移除平台前缀/啤酒小站中缀并修复括号，保留可查证的挂牌名。"""
    text = re.sub(r"\s+", " ", str(shop_name or "")).strip()
    text = text.replace("（", "(").replace("）", ")")
    for _ in range(3):
        text = re.sub(r"^(?:闪购|秒送|自营秒送|自营|连锁)\s*", "", text)
    text = text.replace("-啤酒小站", "").replace("啤酒小站", "")
    # 行政区前缀只在括号内部剥离，避免误删品牌名。
    text = re.sub(r"([\(])[一-鿿]{1,8}区", r"\1", text)
    if text.count("(") > text.count(")"):
        text += ")" * (text.count("(") - text.count(")"))
    # 输出统一使用中文全角括号，避免同一片区因括号形态不同出现重复。
    return text.strip(" -").replace("(", "（").replace(")", "）")


def normalize_shop_name(shop_name: str) -> tuple[str, str]:
    """返回 ``(片区键, 可查证的规范显示名)``。

    后缀按最长优先处理，避免 ``云仓店`` 被错误截成 ``云``。
    """
    cleaned = _clean_shop_name(shop_name)
    match = re.search(r"[（]([^（）]*)[）]", cleaned)
    area = match.group(1).strip() if match else cleaned
    for suffix in ("云仓店", "仓店", "云仓", "店"):
        if area.endswith(suffix):
            area = area[: -len(suffix)]
            break
    area = area.strip(" -") or cleaned
    return area, cleaned


def _normalize_product_name(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("×", "*").replace("x", "*").replace("X", "*")
    return re.sub(r"(?i)ml", "ml", text)


def extract_spec(product_name: str) -> str:
    text = _normalize_product_name(product_name)
    match = re.search(
        r"(\d+\s*ml\s*\*\s*\d+\s*\+\s*\d+\s*[瓶听罐])",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        match = re.search(
            r"(\d+\s*ml\s*\*\s*\d+\s*[瓶听罐])",
            text,
            flags=re.IGNORECASE,
        )
    if not match:
        return ""
    return re.sub(r"\s+", "", match.group(1)).replace("罐", "听")


def _product_family(product_name: str, spec: str) -> str:
    family = _normalize_product_name(product_name)
    if spec:
        family = family.replace(spec, "")
    return re.sub(r"\s+", "", family).lower()


def _province_from_region(region: str) -> str:
    text = str(region or "")
    for province in DEFAULT_1998_PROVINCE_THRESHOLDS:
        if province in text:
            return province
    for city, province in CITY_TO_PROVINCE.items():
        if city in text:
            return province
    return ""


def _product_type(product_name: str) -> str:
    text = product_name.lower()
    if "1998" in text:
        return "1998"
    if "u8" in text or "燕京" in text:
        return "u8"
    return ""


def threshold_for(
    product_name: str,
    spec: str,
    region: str,
    thresholds: Optional[Mapping[str, Mapping[int, float]]] = None,
) -> Optional[float]:
    """按产品、规格和省份返回合格线；未知口径返回 None。"""
    match = re.search(r"\*(\d+)", spec)
    if not match:
        return None
    count = int(match.group(1))
    product_type = _product_type(product_name)
    if not product_type:
        return None
    configured = thresholds or DEFAULT_THRESHOLDS
    if product_type == "1998":
        province = _province_from_region(region)
        province_rules = DEFAULT_1998_PROVINCE_THRESHOLDS.get(province)
        if province_rules and count in province_rules:
            return float(province_rules[count])
    rules = configured.get(product_type, {})
    value = rules.get(count)
    return float(value) if value is not None else None


def _extract_wps_images(xlsx_path: str) -> dict[str, bytes]:
    """读取 WPS cellimages 的 ``ID -> 图片字节`` 映射。"""
    images: dict[str, bytes] = {}
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        names = set(archive.namelist())
        if "xl/cellimages.xml" not in names:
            return images
        cellimages = archive.read("xl/cellimages.xml").decode("utf-8")
        rels = archive.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
        id_to_rid = {
            image_id: rid
            for image_id, rid in re.findall(
                r'name="(ID_[A-Fa-f0-9]+)".*?r:embed="(rId[^" ]+)"',
                cellimages,
                flags=re.S,
            )
        }
        rid_to_target = {
            rid: target
            for rid, target in re.findall(
                r'Id="(rId[^" ]+)".*?Target="([^"]+)"', rels,
                flags=re.S,
            )
        }
        for image_id, rid in id_to_rid.items():
            target = rid_to_target.get(rid, "")
            target = str(PurePosixPath("xl", target))
            if target in names:
                images[image_id] = archive.read(target)
    return images


def _image_id(value: Any) -> str:
    match = re.search(r"(ID_[A-Fa-f0-9]+)", str(value or ""))
    return match.group(1) if match else ""


def parse_source_workbook(xlsx_path: str) -> tuple[list[SourceRecord], list[str]]:
    """读取 A:P 巡查表，返回标准化记录和解析警告。"""
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(xlsx_path)
    images = _extract_wps_images(xlsx_path)
    values_wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    formulas_wb = load_workbook(xlsx_path, data_only=False, read_only=False)
    records: list[SourceRecord] = []
    warnings: list[str] = []
    for sheet_name in values_wb.sheetnames:
        ws = values_wb[sheet_name]
        formula_ws = formulas_wb[sheet_name]
        if ws.max_column < 7:
            continue
        for row in range(3, ws.max_row + 1):
            region = str(ws.cell(row, 2).value or "").strip()
            shop = str(ws.cell(row, 4).value or "").strip()
            if not shop:
                continue
            source_platform = str(ws.cell(row, 3).value or "").strip()
            platform = SOURCE_PLATFORM_ALIASES.get(source_platform, "")
            if not platform:
                warnings.append(f"第{row}行平台无法映射：{source_platform or '空'}")
                continue
            product = _normalize_product_name(ws.cell(row, 5).value)
            no_sale = not product
            spec = extract_spec(product)
            if product and not spec:
                warnings.append(f"第{row}行规格无法解析：{product}")
            final_price = _safe_float(ws.cell(row, 7).value)
            delivery_fee = _safe_float(ws.cell(row, 12).value) or 0.0
            theory_price = _safe_float(ws.cell(row, 13).value)
            if theory_price is None and final_price is not None:
                theory_price = round(final_price - delivery_fee, 2)
            if product and final_price is None:
                warnings.append(f"第{row}行缺少成交价：{shop} / {product}")
            image_value = ws.cell(row, 15).value
            if not image_value:
                image_value = formula_ws.cell(row, 15).value
            image_bytes = images.get(_image_id(image_value))
            key, display = normalize_shop_name(shop)
            records.append(
                SourceRecord(
                    row_number=row,
                    region=region,
                    shop_name=shop,
                    platform=platform,
                    product_name=product,
                    spec=spec,
                    final_price=final_price,
                    theory_price=theory_price,
                    image_bytes=image_bytes,
                    shop_key=key,
                    display_name=display,
                    is_no_sale=no_sale,
                )
            )
    if not records:
        raise ValueError("未读取到有效的巡查表数据")
    return records, warnings


def merge_records(
    records: Iterable[SourceRecord],
    *,
    shop_name_mapping: Optional[Mapping[str, str]] = None,
    no_sale_text: str = DEFAULT_NO_SALE_TEXT,
    thresholds: Optional[Mapping[str, Mapping[int, float]]] = None,
) -> tuple[list[StoreCheckRow], list[str], list[str]]:
    """按片区/规格/平台合并并重算违规状态。"""
    records = list(records)
    pending: list[str] = []
    logs: list[str] = []
    first_shop: OrderedDict[str, str] = OrderedDict()
    first_region: dict[str, str] = {}
    for record in records:
        first_shop.setdefault(record.shop_key, record.display_name)
        old_region = first_region.setdefault(record.shop_key, record.region)
        if old_region and record.region and old_region != record.region:
            pending.append(
                f"片区 {record.shop_key} 出现多个区域：{old_region} / {record.region}"
            )
    def display_for(key: str) -> str:
        if shop_name_mapping and key in shop_name_mapping:
            return str(shop_name_mapping[key])
        return first_shop.get(key, key)

    sold_groups: OrderedDict[tuple[str, str, str], dict[str, SourceRecord]] = (
        OrderedDict()
    )
    no_sale_groups: OrderedDict[str, dict[str, SourceRecord]] = OrderedDict()
    for record in records:
        if record.is_no_sale:
            no_sale_groups.setdefault(record.shop_key, {})[record.platform] = record
            continue
        group_key = (
            record.shop_key,
            _product_family(record.product_name, record.spec),
            record.spec,
        )
        by_platform = sold_groups.setdefault(group_key, {})
        previous = by_platform.get(record.platform)
        # 大纲定义为最低成交价；缺价记录始终低优先级。
        if previous is None or (
            record.final_price is not None
            and (
                previous.final_price is None
                or record.final_price < previous.final_price
            )
        ):
            by_platform[record.platform] = record
        elif previous and record.final_price == previous.final_price:
            logs.append(
                f"同价重复保留首条：{record.shop_name} / {record.product_name} / "
                f"{record.platform}（第{previous.row_number}行）"
            )
    rows: list[StoreCheckRow] = []
    for (shop_key, _family, spec), by_platform in sold_groups.items():
        first = next(iter(by_platform.values()))
        threshold = threshold_for(
            first.product_name, spec, first.region, thresholds
        )
        if threshold is None:
            status = "待核对"
            pending.append(f"未找到合格线：{first.product_name} / {spec}")
        else:
            status = "是" if any(
                item.theory_price is not None and item.theory_price < threshold
                for item in by_platform.values()
            ) else "否"
        rows.append(
            StoreCheckRow(
                shop_name=display_for(shop_key),
                product_name=first.product_name,
                status=status,
                platform_values={platform: by_platform.get(platform) for platform in PLATFORMS},
                shop_key=shop_key,
                spec=spec,
                threshold=threshold,
            )
        )
        logs.append(
            f"合并片区={shop_key} 规格={spec} 平台数={len(by_platform)} 状态={status}"
        )
    for shop_key, by_platform in no_sale_groups.items():
        if any(row.shop_key == shop_key for row in rows):
            continue
        rows.append(
            StoreCheckRow(
                shop_name=display_for(shop_key),
                product_name=no_sale_text,
                status="/",
                platform_values={platform: by_platform.get(platform) for platform in PLATFORMS},
                no_sale=True,
                shop_key=shop_key,
            )
        )
    shop_order = {
        shop_key: min(
            record.row_number
            for record in records
            if record.shop_key == shop_key
        )
        for shop_key in {item.shop_key for item in rows}
    }
    rows.sort(
        key=lambda row: (
            row.no_sale,
            shop_order.get(row.shop_key, 10**9),
            SPEC_ORDER.get(row.spec, 99),
        )
    )
    return rows, pending, logs


def _thin_border() -> Border:
    side = Side(style="thin", color="FFB7B7B7")
    return Border(left=side, right=side, top=side, bottom=side)


def _build_output_workbook(rows: list[StoreCheckRow], input_path: str, pending: list[str], logs: list[str]) -> tuple[Workbook, list[tuple[str, int, int, bytes, str]]]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:F1")
    ws.merge_cells("G1:I1")
    ws.merge_cells("J1:L1")
    headers = [
        ("A1", "门店"),
        ("B1", "产品"),
        ("C1", "是否违规"),
        ("D1", "产品理论成交价格\n（产品成交价格-打包、配送费）"),
        ("G1", "产品实际成交单价\n（最终付款价格）"),
        ("J1", "图片"),
    ]
    for cell_ref, value in headers:
        ws[cell_ref] = value
    for col, value in zip("DEF", PLATFORMS):
        ws[f"{col}2"] = value
    for col, value in zip("GHI", PLATFORMS):
        ws[f"{col}2"] = value
    for col, value in zip("JKL", PLATFORMS):
        ws[f"{col}2"] = value
    header_font = Font(name="微软雅黑", size=11, bold=True)
    data_font = Font(name="宋体", size=11, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_alignment = Alignment(horizontal="center", vertical="center")
    border = _thin_border()
    for row in range(1, 3):
        for col in range(1, 13):
            cell = ws.cell(row, col)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = copy(border)
    ws.row_dimensions[1].height = 31
    ws.row_dimensions[2].height = 20
    widths = {"A": 45, "B": 24.77, "C": 16.06, "J": 11.5}
    for col in "DEFGHIKL":
        widths.setdefault(col, 13)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    image_records: list[tuple[str, int, int, bytes, str]] = []
    for output_row, item in enumerate(rows, start=3):
        ws.row_dimensions[output_row].height = 20
        values = [item.shop_name, item.product_name, item.status]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(output_row, col, value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = copy(border)
        for offset, platform in enumerate(PLATFORMS):
            record = item.platform_values.get(platform)
            theory_cell = ws.cell(output_row, 4 + offset)
            actual_cell = ws.cell(output_row, 7 + offset)
            theory_cell.value = record.theory_price if record and record.theory_price is not None else "/"
            actual_cell.value = record.final_price if record and record.final_price is not None else "/"
            for cell in (theory_cell, actual_cell):
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = copy(border)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0#"
            image_cell = ws.cell(output_row, 10 + offset)
            if record and record.image_bytes:
                placeholder = f"PH_{output_row}_{offset}"
                image_cell.value = f'=_xlfn.DISPIMG("{placeholder}",1)'
                image_records.append(
                    ("Sheet1", output_row, 10 + offset, record.image_bytes, placeholder)
                )
            else:
                image_cell.value = "/"
            image_cell.font = data_font
            image_cell.alignment = data_alignment
            image_cell.border = copy(border)
    log_ws = wb.create_sheet("转换日志")
    log_ws.append(["字段", "内容"])
    log_ws.append(["规则版本", CONVERTER_VERSION])
    log_ws.append(["输入文件", os.path.abspath(input_path)])
    log_ws.append(["生成时间", datetime.now().isoformat(timespec="seconds")])
    log_ws.append(["合并/判定记录", ""])
    for entry in logs:
        log_ws.append(["", entry])
    pending_ws = wb.create_sheet("待确认清单")
    pending_ws.append(["序号", "待确认事项"])
    for index, message in enumerate(pending, start=1):
        pending_ws.append([index, message])
    for extra_ws in (log_ws, pending_ws):
        extra_ws.freeze_panes = "A2"
        extra_ws.column_dimensions["A"].width = 18
        extra_ws.column_dimensions["B"].width = 100
        for row in extra_ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = copy(border)
    if rows:
        ws.freeze_panes = "D3"
        ws.auto_filter.ref = f"A2:L{len(rows) + 2}"
    return wb, image_records


def _inject_wps_cellimages(
    xlsx_path: str,
    image_records: Iterable[tuple[str, int, int, bytes, str]],
) -> None:
    """把图片和占位符注入新工作簿的 WPS cellimages 部件。"""
    records = list(image_records)
    if not records:
        return
    cellimages: list[str] = []
    relationships: list[str] = []
    media: list[tuple[str, bytes]] = []
    placeholder_map: dict[str, str] = {}
    for index, (_sheet, _row, _col, image_bytes, placeholder) in enumerate(records, start=1):
        image_id = f"ID_{index:032X}"
        rid = f"rId{index}"
        filename = f"image_store_{index}.jpeg"
        placeholder_map[placeholder] = image_id
        cellimages.append(
            f'<etc:cellImage><xdr:pic><xdr:nvPicPr>'
            f'<xdr:cNvPr id="{index + 1}" name="{image_id}"/>'
            f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
            f'</xdr:nvPicPr><xdr:blipFill><a:blip r:embed="{rid}"/>'
            f'<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="5715000" cy="12439650"/>'
            f'</a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f'</xdr:spPr></xdr:pic></etc:cellImage>'
        )
        relationships.append(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="media/{filename}"/>'
        )
        media.append((filename, image_bytes))
    cellimages_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<etc:cellImages '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">'
        + "".join(cellimages)
        + "</etc:cellImages>"
    )
    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(relationships)
        + "</Relationships>"
    )
    temporary_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as source, zipfile.ZipFile(
        temporary_path, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            name = item.filename
            if name in {"xl/cellimages.xml", "xl/_rels/cellimages.xml.rels"}:
                continue
            data = source.read(name)
            if name == "[Content_Types].xml":
                text = data.decode("utf-8")
                if 'Extension="jpeg"' not in text:
                    text = text.replace(
                        "</Types>",
                        '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>',
                    )
                text = text.replace(
                    "</Types>",
                    '<Override PartName="/xl/cellimages.xml" '
                    'ContentType="application/vnd.wps-officedocument.cellimage+xml"/>'
                    "</Types>",
                )
                data = text.encode("utf-8")
            elif name == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8")
                if "cellimages.xml" not in text:
                    text = text.replace(
                        "</Relationships>",
                        '<Relationship Id="rId_cellimages" '
                        'Type="http://www.wps.cn/officeDocument/2020/cellImage" '
                        'Target="cellimages.xml"/></Relationships>',
                    )
                data = text.encode("utf-8")
            elif name.startswith("xl/worksheets/") and name.endswith(".xml"):
                text = data.decode("utf-8")
                for placeholder, image_id in placeholder_map.items():
                    text = text.replace(placeholder, image_id)
                data = text.encode("utf-8")
            target.writestr(item, data)
        target.writestr("xl/cellimages.xml", cellimages_xml.encode("utf-8"))
        target.writestr("xl/_rels/cellimages.xml.rels", rels_xml.encode("utf-8"))
        for filename, image_bytes in media:
            target.writestr(f"xl/media/{filename}", image_bytes)
    shutil.move(temporary_path, xlsx_path)


def validate_output_workbook(xlsx_path: str) -> list[str]:
    """校验输出包的 XML、图片引用和媒体闭环。

    返回空列表表示通过；返回的文本可直接写入转换日志或提示用户。
    """
    errors: list[str] = []
    try:
        with zipfile.ZipFile(xlsx_path, "r") as archive:
            bad_name = archive.testzip()
            if bad_name:
                errors.append(f"压缩包校验失败：{bad_name}")
            names = set(archive.namelist())
            sheet_names = [
                name for name in names
                if name.startswith("xl/worksheets/") and name.endswith(".xml")
            ]
            for name in sheet_names:
                try:
                    sheet_xml = archive.read(name).decode("utf-8")
                    # ElementTree 解析能捕获 XML 截断/非法字符。
                    ElementTree.fromstring(sheet_xml)
                except Exception as exc:
                    errors.append(f"{name} XML 无法解析：{exc}")
            if "xl/cellimages.xml" not in names:
                return errors
            cellimages = archive.read("xl/cellimages.xml").decode("utf-8")
            ElementTree.fromstring(cellimages)
            rels = archive.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
            ElementTree.fromstring(rels)
            image_ids = set(re.findall(r'name="(ID_[A-Fa-f0-9]+)"', cellimages))
            rel_map = dict(re.findall(
                r'Id="(rId[^" ]+)".*?Target="([^"]+)"', rels,
                flags=re.S,
            ))
            embed_ids = set(re.findall(r'r:embed="(rId[^" ]+)"', cellimages))
            missing_rids = embed_ids - set(rel_map)
            if missing_rids:
                errors.append(f"cellimages 缺少关系：{sorted(missing_rids)}")
            for target in rel_map.values():
                media_name = str(PurePosixPath("xl", target))
                if media_name not in names:
                    errors.append(f"图片媒体不存在：{media_name}")
            for name in sheet_names:
                sheet_xml = archive.read(name).decode("utf-8")
                referenced = set(re.findall(r'DISPIMG\("(ID_[A-Fa-f0-9]+)"', sheet_xml))
                unknown = referenced - image_ids
                if unknown:
                    errors.append(f"{name} 引用了不存在的图片：{sorted(unknown)}")
    except (OSError, zipfile.BadZipFile) as exc:
        errors.append(f"输出文件无法读取：{exc}")
    return errors


def convert_inspection_to_store_check(
    input_path: str,
    output_dir: Optional[str] = None,
    *,
    output_name: Optional[str] = None,
    shop_name_mapping: Optional[Mapping[str, str]] = None,
    no_sale_text: str = DEFAULT_NO_SALE_TEXT,
    thresholds: Optional[Mapping[str, Mapping[int, float]]] = None,
) -> ConversionResult:
    """将一张巡查表转换为一张门店价格检查表。"""
    records, parse_warnings = parse_source_workbook(input_path)
    rows, pending, logs = merge_records(
        records,
        shop_name_mapping=shop_name_mapping,
        no_sale_text=no_sale_text,
        thresholds=thresholds,
    )
    pending = parse_warnings + pending
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(input_path))
    os.makedirs(output_dir, exist_ok=True)
    if not output_name:
        city = (records[0].region or "").replace("市", "") or "门店"
        date_match = re.search(r"(20\d{6})", os.path.basename(input_path))
        date_text = date_match.group(1) if date_match else datetime.now().strftime("%Y%m%d")
        output_name = f"{city}门店价格检查表_{date_text}.xlsx"
    output_path = os.path.join(output_dir, output_name)
    workbook, image_records = _build_output_workbook(
        rows, input_path, pending, logs
    )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output_path)
    _inject_wps_cellimages(output_path, image_records)
    validation_errors = validate_output_workbook(output_path)
    if validation_errors:
        raise ValueError("输出文件校验失败：" + "；".join(validation_errors))
    return ConversionResult(output_path, rows, pending, logs)


# 便于 UI 和旧代码使用的中文别名。
convert巡查表到门店价格检查表 = convert_inspection_to_store_check


__all__ = [
    "CONVERTER_VERSION",
    "ConversionResult",
    "DEFAULT_NO_SALE_TEXT",
    "PLATFORMS",
    "SourceRecord",
    "StoreCheckRow",
    "convert_inspection_to_store_check",
    "convert巡查表到门店价格检查表",
    "extract_spec",
    "merge_records",
    "normalize_shop_name",
    "parse_source_workbook",
    "threshold_for",
    "validate_output_workbook",
]
