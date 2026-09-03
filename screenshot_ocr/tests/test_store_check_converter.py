import os
import re
import sys
import tempfile
import unittest
import zipfile

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from store_check_converter import (  # noqa: E402
    SourceRecord,
    convert_inspection_to_store_check,
    extract_spec,
    filter_u8_records,
    merge_records,
    normalize_shop_name,
    parse_source_workbook,
    threshold_for,
    validate_output_workbook,
)


class StoreCheckConverterTest(unittest.TestCase):
    def test_shop_suffix_longest_first_and_parentheses(self):
        key, display = normalize_shop_name("秒送 京东酒世界（富春花园云仓店）")
        self.assertEqual(key, "富春花园")
        self.assertEqual(display, "京东酒世界（富春花园云仓店）")
        key, display = normalize_shop_name("京东酒世界（滇池度假区仓店）")
        self.assertEqual(key, "滇池度假区")
        self.assertEqual(display, "京东酒世界（滇池度假区仓店）")

    def test_spec_normalization(self):
        self.assertEqual(extract_spec("燕京U8 500ml×12罐"), "500ml*12听")
        self.assertEqual(extract_spec("漓泉1998 500ml*9+3听"), "500ml*9+3听")

    def test_merge_uses_lowest_price_and_recalculates_status(self):
        def record(row, platform, price, theory):
            return SourceRecord(
                row_number=row,
                region="昆明市",
                shop_name="京东酒世界（富春花园仓店）",
                platform=platform,
                product_name="燕京U8 500ml*6瓶",
                spec="500ml*6瓶",
                final_price=price,
                theory_price=theory,
                shop_key="富春花园",
                display_name="京东酒世界（富春花园仓店）",
            )

        rows, pending, _ = merge_records([
            record(1, "美团闪购", 40, 36),
            record(2, "美团闪购", 38, 28),
            record(3, "京东闪送", 30, 30),
        ])
        self.assertFalse(pending)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].platform_values["美团闪购"].final_price, 38)
        self.assertEqual(rows[0].status, "是")

    def test_different_chains_same_area_are_not_merged(self):
        records = []
        for row, shop in enumerate(
            ("京东酒世界（南亚店）", "另一连锁（南亚店）"), start=1
        ):
            area, display = normalize_shop_name(shop)
            records.append(
                SourceRecord(
                    row_number=row,
                    region="昆明市",
                    shop_name=shop,
                    platform="美团闪购",
                    product_name="燕京U8 500ml*6瓶",
                    spec="500ml*6瓶",
                    final_price=30,
                    theory_price=30,
                    shop_key=f"{display.split('（', 1)[0]}|{area}",
                    display_name=display,
                )
            )
        rows, pending, _ = merge_records(records)
        self.assertFalse(pending)
        self.assertEqual(len(rows), 2)

    def test_same_shop_key_in_different_cities_stays_separate(self):
        records = []
        for row, city in enumerate(("昆明市", "贵阳市"), start=1):
            records.append(
                SourceRecord(
                    row_number=row,
                    region=city,
                    shop_name="京东酒世界（中心店）",
                    platform="美团闪购",
                    product_name="燕京U8 500ml*6瓶",
                    spec="500ml*6瓶",
                    final_price=30,
                    theory_price=30,
                    shop_key="京东酒世界|中心",
                    display_name="京东酒世界（中心店）",
                )
            )
        rows, pending, _ = merge_records(records)
        self.assertFalse(pending)
        self.assertEqual(len(rows), 2)
        self.assertEqual({row.city for row in rows}, {"昆明市", "贵阳市"})

    def test_conversion_scope_is_u8_only(self):
        base = dict(
            row_number=1,
            region="昆明市",
            shop_name="测试门店（南亚店）",
            platform="美团闪购",
            spec="500ml*6瓶",
            final_price=30,
            theory_price=30,
            shop_key="测试门店|南亚",
            display_name="测试门店（南亚店）",
        )
        u8 = SourceRecord(product_name="燕京U8 500ml*6瓶", **base)
        other = SourceRecord(
            product_name="漓泉1998 500ml*6瓶", row_number=2, **{k: v for k, v in base.items() if k != "row_number"}
        )
        kept, skipped = filter_u8_records([u8, other])
        self.assertEqual(kept, [u8])
        self.assertEqual(len(skipped), 1)

    def test_1998_province_thresholds(self):
        self.assertEqual(
            threshold_for("漓泉1998 500ml*12瓶", "500ml*12瓶", "广东省"),
            70,
        )
        self.assertEqual(
            threshold_for("漓泉1998 500ml*12瓶", "500ml*12瓶", "广西南宁市"),
            60,
        )
        self.assertEqual(
            threshold_for("漓泉1998 500ml*6瓶", "500ml*6瓶", "广东省"),
            29.9,
        )
        self.assertIsNone(
            threshold_for("漓泉1998 500ml*12瓶", "500ml*12瓶", "未知市")
        )

    def test_real_workbook_parse_and_output_integrity(self):
        root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        source = os.path.join(
            root,
            "output",
            "巡查表_20260902_143723",
            "价格巡查表_20260902_143723.xlsx",
        )
        records, warnings = parse_source_workbook(source)
        self.assertEqual(len(records), 43)
        self.assertFalse(warnings)
        self.assertTrue(all(item.image_bytes for item in records))
        with tempfile.TemporaryDirectory() as output_dir:
            result = convert_inspection_to_store_check(
                source,
                output_dir=output_dir,
                output_name="转换测试.xlsx",
            )
            self.assertEqual(result.output_path, os.path.join(output_dir, "转换测试.xlsx"))
            self.assertTrue(os.path.isfile(result.output_path))
            self.assertEqual(validate_output_workbook(result.output_path), [])
            with zipfile.ZipFile(result.output_path) as archive:
                self.assertIn("xl/cellimages.xml", archive.namelist())
                self.assertIn("xl/_rels/cellimages.xml.rels", archive.namelist())
                sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
                ids = set(re.findall(r'"(ID_[A-F0-9]{32})"', sheet_xml))
                cellimages = archive.read("xl/cellimages.xml").decode("utf-8")
                self.assertTrue(ids)
                for image_id in ids:
                    self.assertIn(image_id, cellimages)
                self.assertNotIn("#REF!", sheet_xml)

    def test_output_includes_city_column_and_hq_detail_name(self):
        root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        source = os.path.join(
            root,
            "output",
            "巡查表_20260902_171538",
            "价格巡查表_20260902_171538.xlsx",
        )
        with tempfile.TemporaryDirectory() as output_dir:
            result = convert_inspection_to_store_check(source, output_dir=output_dir)
            self.assertEqual(
                os.path.basename(result.output_path),
                "总部供货渠道价格明细_20260902.xlsx",
            )
            wb = load_workbook(result.output_path, data_only=False)
            self.assertEqual(wb.properties.title, "总部供货渠道价格明细")
            ws = wb["总部供货渠道价格明细"]
            self.assertEqual(ws.max_column, 13)
            self.assertEqual(
                [ws.cell(1, column).value for column in range(1, 5)],
                ["城市", "门店", "产品", "是否违规"],
            )
            self.assertEqual(ws.freeze_panes, "E3")
            self.assertTrue(ws.auto_filter.ref.startswith("A2:M"))
            cities = {
                ws.cell(row, 1).value
                for row in range(3, ws.max_row + 1)
                if ws.cell(row, 1).value
            }
            self.assertTrue({"昆明市", "贵阳市", "遵义市"}.issubset(cities))


if __name__ == "__main__":
    unittest.main()
