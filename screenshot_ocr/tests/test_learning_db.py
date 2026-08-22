"""
店铺/城市智能匹配数据库 — 测试套件

覆盖：标准化 / L1-L4 分级匹配 / 模糊候选不写库 / 学习闭环 / 投喂幂等 / 冲突检测与裁决 / 迁移幂等

运行（使用独立测试库，不影响真实 data/ocr_learning.db）：
    OCR_LEARNING_DB=/tmp/test_learning.db python3 tests/test_learning_db.py
"""

import os
import sys
import tempfile

# 必须在 import database 之前设置独立测试库
TEST_DB = os.environ.get("OCR_LEARNING_DB") or os.path.join(tempfile.gettempdir(), "test_ocr_learning.db")
if os.path.exists(TEST_DB):
    os.remove(TEST_DB)
os.environ["OCR_LEARNING_DB"] = TEST_DB

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import database
from database.matcher import normalize, match
from database import importer

PASS = 0
FAIL = 0


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  ✅ {name}")
    else:
        FAIL += 1
        print(f"  ❌ {name}")


def test_normalize():
    print("== 标准化 ==")
    check("去空格", normalize("广州 天河 XX 超市") == "广州天河XX超市")
    check("全角括号统一", normalize("百佳汇生活超市（玉州店）") == "百佳汇生活超市(玉州店)")
    check("全角数字转半角", normalize("５００ml") == "500ml")
    check("OCR错别字 广卅->广州", "广州" in normalize("广卅天河XX超市"))
    check("尾部噪声清理", normalize("美宜佳（商品街店）＞") == "美宜佳(商品街店)")


def test_match_levels():
    print("== 分级匹配 ==")
    # L1 canonical
    r = match("测试旗舰店", record_history=False)
    check("新店无命中 L6", r.level == 6)

    database.learn_correction("测试旗舰店", "测试旗舰店", "广州市", operator="test")
    r = match("测试旗舰店", record_history=False)
    check("L1 canonical 命中", r.level == 1 and r.city == "广州市")

    # L2 alias：OCR 变体
    database.learn_correction("广卅天河XX超市", "广州天河XX超市", "广州市", operator="test")
    r = match("广卅天河XX超市", record_history=False)
    check("L2 alias 命中(OCR变体归并)", r.level == 2 and r.shop_id is not None)

    # L3 标准化：加空格变体
    database.learn_correction("百佳汇生活超市（玉州店）", "百佳汇生活超市（玉州店）", "三亚市", operator="test")
    r = match("百佳汇生活超市 (玉州店) ", record_history=False)
    check("L3 标准化命中", r.level == 3)

    # L4 历史修正记录（learn 同时登记了别名，可能走 L2 命中，结果一致）
    database.learn_correction("雀嘻嘻入口店", "惠到家（远东店）", "海口市", operator="test")
    r = match("雀嘻嘻入口店", record_history=False)
    check("历史修正命中(别名/修正记录)", r.level in (2, 4) and r.canonical_name == "惠到家（远东店）")

    # L5 模糊：不写 canonical
    database.learn_correction("老王便利店（滨江路店）", "老王便利店（滨江路店）", "广州市", operator="test")
    r = match("老王便利店(滨江路店)", record_history=False)
    check("L3 命中(标点差异)", r.level == 3)
    r = match("老王便利店滨江路店", record_history=False)
    check("L5 高相似候选(不自动写库)", r.level == 5 and r.candidates)


def test_learn_loop():
    print("== 学习闭环（修正一次，下次命中） ==")
    # 第一次：无法确认
    r = match("佰佳汇生活超市（玉州店）", record_history=False)
    check("第一次无命中", r.level == 6)
    # 人工修正
    database.learn_correction("佰佳汇生活超市（玉州店）", "百佳汇生活超市（玉州店）", "三亚市", operator="test")
    # 第二次：直接命中
    r = match("佰佳汇生活超市（玉州店）", record_history=False)
    check("第二次命中且城市正确", r.level in (2, 4) and r.city == "三亚市")


def test_import_idempotent():
    print("== 投喂幂等 ==")
    wb_path = os.path.join(tempfile.gettempdir(), "feed_idem.xlsx")
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["分公司", "区域", "平台", "店铺", "产品", "原价"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["", "南宁市", "美团", "测试超市（青秀店）", "燕京U8", "60"])
    ws.append(["", "南宁市", "美团", "测试超市（青秀店）", "燕京U8", "65"])  # 同行重复
    wb.save(wb_path)

    r1 = database.import_excel(wb_path, operator="test")
    check("首次投喂", r1["status"] == "ok" and r1["new_shops"] == 1)
    check("修正记录关联批次", any(c["batch_id"] == r1["batch_id"] for c in database.list_corrections()))
    before = database.get_stats()
    r2 = database.import_excel(wb_path, operator="test")
    check("重复投喂识别为重复", r2["duplicate"] is True)
    after = database.get_stats()
    check("重复投喂不新增数据",
          after["shops"] == before["shops"] and after["aliases"] == before["aliases"])


def test_import_rejects_summary_sheets_and_invalid_shop_names():
    print("== 投喂表头与内容校验 ==")
    wb_path = os.path.join(tempfile.gettempdir(), "feed_sheet_guard.xlsx")
    from openpyxl import Workbook

    wb = Workbook()
    detail = wb.active
    detail.title = "巡查明细"
    detail.append(["分公司", "所属主要区域", "平台", "店铺名称"])
    detail.append(["", "所属主要区域", "", "店铺名称"])
    detail.append(["", "广州市", "美团", "正常便利店"])
    detail.append(["", "广州市", "美团", "500ml*12瓶"])
    summary = wb.create_sheet("汇总")
    summary.append(["汇总"])
    summary.append(["序号", "规格", "合格线(元)", "产品规格"])
    summary.append([1, "500ml*12瓶", 60, "500ml*12瓶"])
    wb.save(wb_path)

    report = database.import_excel(wb_path, operator="test")
    check("只导入明细表中的有效店铺", report["total_rows"] == 1)
    check("规格值被忽略", report["ignored_rows"] == 1)
    check("汇总表被跳过", report["detail"]["skipped_sheets"] == ["汇总"])
    check("正常店铺仍可匹配", database.get_shop_city("正常便利店")["city"] == "广州市")


def test_cleanup_invalid_imports():
    print("== 异常投喂清理 ==")
    invalid_id = database.repository.upsert_shop("500ml*12瓶", source="import")
    normal_id = database.repository.upsert_shop("保留的正常店铺", source="import")
    candidates = database.list_invalid_imported_shops()
    candidate_ids = {row["shop_id"] for row in candidates}
    check("异常规格出现在清理预览", invalid_id in candidate_ids)
    check("正常店铺不出现在清理预览", normal_id not in candidate_ids)
    deleted = database.delete_invalid_imported_shops([invalid_id, normal_id])
    check("只删除异常投喂项", deleted == ["500ml*12瓶"])
    check("正常投喂店铺保留", database.repository.get_shop(normal_id) is not None)


def test_conflict():
    print("== 冲突检测与裁决 ==")
    database.learn_correction("双城超市", "双城超市", "广州市", operator="test")
    out = database.learn_correction("双城超市", "双城超市", "佛山市", operator="test")
    check("同店不同城市 -> 冲突", out["conflict"] is True)
    conflicts = database.get_conflicts()
    check("冲突列表包含该店", any(c["canonical_name"] == "双城超市" for c in conflicts))
    # 裁决
    shop = next(c for c in conflicts if c["canonical_name"] == "双城超市")
    database.resolve_conflict(shop["shop_id"], "广州市", operator="test")
    r = match("双城超市", record_history=False)
    check("裁决后命中广州", r.city == "广州市" and r.is_conflict is False)


def test_alias_conflict_is_visible():
    print("== 别名冲突可见 ==")
    database.learn_correction("相同 OCR 店名", "甲店", "广州市", operator="test")
    database.learn_correction("相同 OCR 店名", "乙店", "佛山市", operator="test")
    rows = database.list_shops(limit=10000)
    names = {row["canonical_name"]: row["status"] for row in rows}
    check("不同店铺共用别名 -> 冲突", names.get("甲店") == "conflict" and names.get("乙店") == "conflict")


def test_migrate_idempotent():
    print("== 迁移幂等 ==")
    from database.migrate import MIGRATE_MARK
    from database.schema import get_meta
    check("迁移标记已存在", get_meta(MIGRATE_MARK) is not None)
    r = database.migrate_shop_city_db()
    check("重复迁移直接跳过", r.get("migrated") is False)


if __name__ == "__main__":
    test_normalize()
    test_match_levels()
    test_learn_loop()
    test_import_idempotent()
    test_import_rejects_summary_sheets_and_invalid_shop_names()
    test_cleanup_invalid_imports()
    test_conflict()
    test_alias_conflict_is_visible()
    test_migrate_idempotent()
    print(f"\n结果: {PASS} 通过, {FAIL} 失败")
    sys.exit(1 if FAIL else 0)
