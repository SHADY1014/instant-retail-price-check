"""Regression tests for readable summary-report headers."""

import os
import sys
import unittest

from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from summary_generator import (
    _build_city_summary,
    _build_province_summary,
    _get_province,
    _get_qualification_line,
    _get_secondary_threshold,
)


RECORDS = [{
    "province": "广东",
    "region": "广州市",
    "product_name": "燕京U8 500ml*12瓶",
    "spec": "500ml*12瓶",
    "qual_line": 60,
    "theory_price": 65,
    "passed": True,
}]


class SummaryLayoutTests(unittest.TestCase):

    def test_explicit_province_names_are_supported(self):
        self.assertEqual(_get_province("广东省"), "广东")
        self.assertEqual(_get_province("广西-南宁"), "广西")

    def test_1998_thresholds_are_province_specific(self):
        self.assertEqual(
            _get_qualification_line("漓泉1998啤酒", "500ml*12瓶", "广东"), 70
        )
        self.assertEqual(
            _get_secondary_threshold("漓泉1998啤酒", "广东", "500ml*12瓶"), 65
        )
        self.assertEqual(
            _get_qualification_line("漓泉1998啤酒", "500ml*12听", "广西"), 60
        )
        self.assertEqual(
            _get_secondary_threshold("漓泉1998啤酒", "广西", "500ml*12听"), 55
        )

    def test_mixed_1998_summary_uses_each_province_rule(self):
        records = [
            {
                "province": "广东", "region": "广州市",
                "product_name": "漓泉1998啤酒 500ml*12瓶",
                "spec": "500ml*12瓶", "theory_price": 69,
                "qual_line": 74.99, "passed": True,
            },
            {
                "province": "广西", "region": "南宁市",
                "product_name": "漓泉1998啤酒 500ml*12瓶",
                "spec": "500ml*12瓶", "theory_price": 59,
                "qual_line": 74.99, "passed": True,
            },
        ]
        workbook = Workbook()
        sheet = workbook.active
        _build_province_summary(sheet, records, start_row=1)

        self.assertEqual(sheet.cell(3, 6).value, "合格数\n（各省标准以上）")
        self.assertEqual(sheet.cell(4, 4).value, 70)
        self.assertEqual(sheet.cell(5, 4).value, 60)
        self.assertEqual(sheet.cell(4, 6).value, 0)
        self.assertEqual(sheet.cell(5, 6).value, 0)
        self.assertNotIn("各省标准元", sheet.cell(3, 9).value)
        self.assertEqual(sheet.cell(3, 9).value, "各省标准以上\n价格")
        self.assertIn("广东", sheet.cell(2, 1).value)
        self.assertIn("≥70元", sheet.cell(2, 1).value)
        self.assertIn("广西", sheet.cell(2, 1).value)
        self.assertIn("≥60元", sheet.cell(2, 1).value)

    def test_summary_headers_are_explicitly_wrapped_and_tall_enough(self):
        workbook = Workbook()
        sheet = workbook.active

        city_start = _build_province_summary(sheet, RECORDS, start_row=1)
        _build_city_summary(sheet, RECORDS, start_row=city_start)

        province_header = 3
        city_header = city_start + 2
        self.assertEqual(sheet.cell(province_header, 6).value, "合格数\n（60元以上）")
        self.assertEqual(sheet.cell(city_header, 7).value, "合格数\n（60元以上）")
        self.assertTrue(sheet.cell(province_header, 6).alignment.wrap_text)
        self.assertTrue(sheet.cell(city_header, 7).alignment.wrap_text)
        self.assertEqual(sheet.cell(province_header, 6).font.color.type, "theme")
        self.assertEqual(sheet.cell(province_header, 6).font.color.theme, 0)
        self.assertEqual(sheet.cell(city_header, 7).font.color.type, "theme")
        self.assertEqual(sheet.cell(city_header, 7).font.color.theme, 0)
        self.assertTrue(sheet.cell(2, 1).alignment.wrap_text)
        self.assertTrue(sheet.cell(city_start + 1, 1).alignment.wrap_text)
        self.assertEqual(sheet.row_dimensions[province_header].height, 42)
        self.assertEqual(sheet.row_dimensions[city_header].height, 42)
        self.assertIn(
            f"A{city_start + 1}:O{city_start + 1}",
            {str(cell_range) for cell_range in sheet.merged_cells.ranges},
        )


if __name__ == "__main__":
    unittest.main()
