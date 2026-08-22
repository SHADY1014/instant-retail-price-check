"""
Excel 写入器 — 保留模板格式，写入数据行 + 公式 + 图片

关键：
  - 不修改原始模板，每次复制新副本输出
  - 保留合并单元格、表头样式、列宽、行高、边框
  - M/N 列写公式 =G{row}-L{row} 和 =G{row}+J{row}+K{row}-L{row}
  - O 列插入截图缩略图
  - A 列（分公司）留空
"""

import io
import os
import shutil
import uuid
import zipfile
from copy import copy
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# 模板路径（与代码同目录，避免用户误操作）
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "模板.xlsx",
)

# 输出目录
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "output",
)

# Sheet 名称映射（模板已删除竞品子表，只保留燕京）
SHEET_NAMES = [
    "1.燕京即时零售渠道价格巡查表",
]

# 数据行起始行号（第1行标题，第2行表头，第3行开始数据）
DATA_START_ROW = 3

# 列映射 (A~P)
COL_MAP = {
    "branch_company": 1,   # A
    "region": 2,           # B
    "platform": 3,         # C
    "shop_name": 4,        # D
    "product_name": 5,     # E
    "original_price": 6,   # F
    "final_price": 7,      # G
    "shop_discount": 8,    # H
    "full_reduction": 9,   # I
    "coupon": 10,          # J
    "red_packet": 11,      # K
    "delivery_fee": 12,    # L
    # M=13 (公式), N=14 (公式), O=15 (图片), P=16 (备注)
}


def _get_data_style(wb, sheet_name):
    """从模板中提取数据行样式（复制第3行的样式作为模板）"""
    ws = wb[sheet_name]
    style = {}
    for col_idx in range(1, 17):
        cell = ws.cell(row=DATA_START_ROW, column=col_idx)
        style[col_idx] = {
            "font": copy(cell.font),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "number_format": cell.number_format,
        }
    return style


def _apply_style(cell, style_dict):
    """应用保存的样式到单元格"""
    cell.font = copy(style_dict["font"])
    cell.alignment = copy(style_dict["alignment"])
    cell.border = copy(style_dict["border"])
    cell.fill = copy(style_dict["fill"])
    cell.number_format = style_dict["number_format"]


def _enable_formula_recalculation(wb):
    """要求 Excel/WPS 打开文件时重新计算所有公式。"""
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


# 压缩图片缓存：{path: (img_bytes, w_px, h_px)}
# 同一张图片在巡查表/汇总表生成中只读取、解码、压缩一次
_COMPRESSED_CACHE = {}
_COMPRESS_MAX_WIDTH = 600   # 巡查表 O 列显示尺寸约 70x140px，600px 已有 4 倍余量
_COMPRESS_QUALITY = 70


def _get_compressed_image(image_path):
    """
    获取压缩后的图片字节与像素尺寸（带缓存）。

    手机截图原始 1080x2400（~0.9MB/张）缩放至宽 600px 并以 quality=70 重编码，
    体积约降 90%，表格内嵌图视觉无差别；OCR 与字段解析不经过本函数，不受影响。

    Returns:
        (img_bytes, w_px, h_px)，失败时回退原图字节
    """
    cached = _COMPRESSED_CACHE.get(image_path)
    if cached:
        return cached

    try:
        with open(image_path, "rb") as f:
            raw = f.read()
    except OSError:
        return b"", 200, 400

    result = (raw, 200, 400)
    try:
        pil = PILImage.open(image_path)
        w_px, h_px = pil.size
        if w_px > _COMPRESS_MAX_WIDTH:
            pil = pil.resize(
                (_COMPRESS_MAX_WIDTH, int(h_px * _COMPRESS_MAX_WIDTH / w_px)),
                PILImage.LANCZOS,
            )
        if pil.mode != "RGB":
            pil = pil.convert("RGB")
        buf = io.BytesIO()
        pil.save(buf, "JPEG", quality=_COMPRESS_QUALITY)
        result = (buf.getvalue(), pil.size[0], pil.size[1])
    except Exception:
        # 解码/压缩失败则回退原图
        try:
            w_px, h_px = PILImage.open(image_path).size
            result = (raw, w_px, h_px)
        except Exception:
            pass

    _COMPRESSED_CACHE[image_path] = result
    return result


def _get_image_ext(image_path):
    """获取图片扩展名（不含点）"""
    ext = os.path.splitext(image_path)[1].lower()
    if ext == ".jpg":
        ext = ".jpeg"
    return ext.lstrip(".")


def _inject_wps_cellimages(xlsx_path, image_records):
    """
    向 xlsx 文件注入 WPS 格式的嵌入图片（cellimages）
    使图片真正嵌入单元格内，与用户手动在 WPS 中插入的图片效果一致。

    Args:
        xlsx_path: xlsx 文件路径
        image_records: list of (row_number, image_path, placeholder) 元组，
                       表示在第 row 行的 O 列插入 image_path 对应的图片
    """
    if not image_records:
        return

    # 并行预压缩所有图片（LANCZOS 缩放较慢，串行约 25ms/张）
    # 预取进缓存后，主循环直接命中，不重复解码
    from concurrent.futures import ThreadPoolExecutor
    _paths = [img_path for _, img_path, _ in image_records]
    with ThreadPoolExecutor(max_workers=8) as _ex:
        list(_ex.map(_get_compressed_image, _paths))

    # 生成每个图片的 ID 和 cellimage XML
    cellimage_xmls = []
    rel_entries = []  # (rId, target)
    media_entries = []  # (media_filename, image_bytes, ext)
    placeholder_map = {}  # placeholder -> img_id

    for i, (row, image_path, placeholder) in enumerate(image_records):
        img_id = f"ID_{uuid.uuid4().hex.upper()}"
        placeholder_map[placeholder] = img_id
        rid = f"rId{i + 1}"
        ext = _get_image_ext(image_path)
        media_filename = f"image_new_{i + 1}.{ext}"
        # 压缩图片（带缓存），并同时取得尺寸，避免重复解码原图
        img_bytes, w_px, h_px = _get_compressed_image(image_path)

        cx = w_px * 9525
        cy = h_px * 9525

        # cellimage XML 节点
        cellimage_xmls.append(
            f'<etc:cellImage><xdr:pic>'
            f'<xdr:nvPicPr><xdr:cNvPr id="{i + 2}" name="{img_id}"/>'
            f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
            f'</xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="{rid}"/>'
            f'<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
            f'</xdr:pic></etc:cellImage>'
        )

        rel_entries.append((rid, f"media/{media_filename}"))
        media_entries.append((media_filename, img_bytes, ext))

    # 构建 cellimages.xml
    cellimages_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<etc:cellImages '
        'xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">'
        + ''.join(cellimage_xmls) +
        '</etc:cellImages>'
    )

    # 构建 cellimages.xml.rels
    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + ''.join(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="{target}"/>'
            for rid, target in rel_entries
        )
        + '</Relationships>'
    )

    # 读取原始 xlsx，重新打包
    tmp_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            # 复制原有文件，跳过需要修改的
            existing_names = set(zin.namelist())
            for item in zin.infolist():
                name = item.filename

                # 跳过已有的 cellimages（如果有）
                if name == "xl/cellimages.xml":
                    continue
                if name == "xl/_rels/cellimages.xml.rels":
                    continue

                data = zin.read(name)

                # 修改 [Content_Types].xml - 添加 cellimages 类型声明
                if name == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    if "cellimage" not in text:
                        # 在 </Types> 前插入
                        text = text.replace(
                            "</Types>",
                            '<Override PartName="/xl/cellimages.xml" '
                            'ContentType="application/vnd.wps-officedocument.cellimage+xml"/>'
                            "</Types>",
                        )
                    data = text.encode("utf-8")

                # 修改 workbook.xml.rels - 添加 cellimages 关系
                if name == "xl/_rels/workbook.xml.rels":
                    text = data.decode("utf-8")
                    if "cellimages" not in text:
                        text = text.replace(
                            "</Relationships>",
                            f'<Relationship Id="rId_cellimages" '
                            f'Type="http://www.wps.cn/officeDocument/2020/cellImage" '
                            f'Target="cellimages.xml"/>'
                            f"</Relationships>",
                        )
                    data = text.encode("utf-8")

                # 替换 sharedStrings.xml 中的占位符为真实图片 ID
                if name == "xl/sharedStrings.xml":
                    text = data.decode("utf-8")
                    # 按占位符长度从长到短排序，避免短占位符误替换长占位符的子串
                    for placeholder, img_id in sorted(placeholder_map.items(), key=lambda x: -len(x[0])):
                        text = text.replace(placeholder, img_id)
                    data = text.encode("utf-8")

                # 也检查 worksheets 中的内联字符串
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    text = data.decode("utf-8")
                    for placeholder, img_id in sorted(placeholder_map.items(), key=lambda x: -len(x[0])):
                        text = text.replace(placeholder, img_id)
                    data = text.encode("utf-8")

                zout.writestr(item, data)

            # 写入新文件（图片已压缩，用 STORED 避免无收益的二次 deflate）
            zout.writestr("xl/cellimages.xml", cellimages_xml.encode("utf-8"))
            zout.writestr("xl/_rels/cellimages.xml.rels", rels_xml.encode("utf-8"))
            for filename, img_bytes, ext in media_entries:
                zi = zipfile.ZipInfo(f"xl/media/{filename}")
                zi.compress_type = zipfile.ZIP_STORED
                zout.writestr(zi, img_bytes)

    shutil.move(tmp_path, xlsx_path)


def write_records(records, sheet_index=0, output_name=None, clear_existing=True):
    """
    将识别结果写入 Excel

    Args:
        records: list[dict]，每个 dict 包含 FormFields.to_dict() 的字段 + "image_path"
        sheet_index: Sheet 索引 (0=燕京, 1=雪花, 2=青岛, 3=百威)
        output_name: 输出文件名（不含路径），None 则自动生成
        clear_existing: 是否清空该 sheet 已有的数据行（保留表头/格式），默认 True

    Returns:
        str: 输出文件路径
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_PATH}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 生成输出文件名
    if not output_name:
        output_name = f"价格巡查表_{timestamp}.xlsx"

    # 每次导出新建子文件夹，便于同批次巡查表+汇总表集中管理
    session_dir = os.path.join(OUTPUT_DIR, f"巡查表_{timestamp}")
    os.makedirs(session_dir, exist_ok=True)

    output_path = os.path.join(session_dir, output_name)

    # 复制模板
    shutil.copy2(TEMPLATE_PATH, output_path)

    # 加载副本
    wb = load_workbook(output_path)
    sheet_name = SHEET_NAMES[sheet_index]
    ws = wb[sheet_name]

    # 提取数据行样式
    style = _get_data_style(wb, sheet_name)

    # 清空已有数据行（保留表头行1~2）
    if clear_existing:
        _clear_data_rows(ws)

    # 从第3行开始写入
    current_row = DATA_START_ROW
    image_records = []  # (row_number, image_path) 用于后续注入 WPS cellimages

    for record in records:
        row = current_row

        # 写入 A~L 列数据
        for field_name, col_idx in COL_MAP.items():
            cell = ws.cell(row=row, column=col_idx)
            value = record.get(field_name, "")
            if value == "" or value is None:
                cell.value = None
            else:
                cell.value = value
            _apply_style(cell, style[col_idx])

        # M 列: 公式 =G{row}-L{row}
        m_cell = ws.cell(row=row, column=13)
        m_cell.value = f"=G{row}-L{row}"
        _apply_style(m_cell, style[13])

        # N 列: 公式 =G{row}+J{row}+K{row}-L{row}
        n_cell = ws.cell(row=row, column=14)
        n_cell.value = f"=G{row}+J{row}+K{row}-L{row}"
        _apply_style(n_cell, style[14])

        # O 列: 写入 WPS DISPIMG 公式（图片 ID 在注入时生成）
        o_cell = ws.cell(row=row, column=15)
        _apply_style(o_cell, style[15])

        image_path = record.get("image_path")
        if image_path and os.path.exists(image_path):
            # 使用 UUID 作为占位符，避免行号冲突
            placeholder = f"PH_{uuid.uuid4().hex.upper()}"
            o_cell.value = f'=_xlfn.DISPIMG("{placeholder}",1)'
            image_records.append((row, image_path, placeholder))
        else:
            o_cell.value = "(无图片)"

        # P 列: 备注
        p_cell = ws.cell(row=row, column=16)
        p_cell.value = record.get("remark", "")
        _apply_style(p_cell, style[16])

        current_row += 1

    # 保存 openpyxl 结果
    _enable_formula_recalculation(wb)
    wb.save(output_path)

    # 注入 WPS 格式的嵌入图片
    if image_records:
        _inject_wps_cellimages(output_path, image_records)

    return output_path


def _clear_data_rows(ws, start_row=DATA_START_ROW):
    """
    清空 sheet 中从 start_row 开始的所有数据行
    保留表头（第1~2行）和格式
    """
    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(start_row, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None


def write_all_brands(records, output_name=None):
    """
    将所有识别结果写入巡查表（模板已删除竞品子表，只有1个燕京Sheet）

    Args:
        records: list[dict]，每个 dict 包含 FormFields.to_dict() 的字段 + "image_path"
        output_name: 输出文件名

    Returns:
        str: 输出文件路径
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"模板文件不存在: {TEMPLATE_PATH}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if not output_name:
        output_name = f"价格巡查表_{timestamp}.xlsx"

    # 每次导出新建子文件夹，便于同批次巡查表+汇总表集中管理
    session_dir = os.path.join(OUTPUT_DIR, f"巡查表_{timestamp}")
    os.makedirs(session_dir, exist_ok=True)

    output_path = os.path.join(session_dir, output_name)

    # 复制模板
    shutil.copy2(TEMPLATE_PATH, output_path)

    # 加载副本
    wb = load_workbook(output_path)

    # 所有记录写入唯一的 Sheet
    all_image_records = []

    for sheet_idx, sheet_name in enumerate(SHEET_NAMES):
        ws = wb[sheet_name]
        style = _get_data_style(wb, sheet_name)

        # 清空已有数据行
        _clear_data_rows(ws)

        # 从第3行开始写入
        current_row = DATA_START_ROW

        for record in records:
            row = current_row

            # 写入 A~L 列数据
            for field_name, col_idx in COL_MAP.items():
                cell = ws.cell(row=row, column=col_idx)
                value = record.get(field_name, "")
                if value == "" or value is None:
                    cell.value = None
                else:
                    cell.value = value
                _apply_style(cell, style[col_idx])

            # M 列: 公式 =G{row}-L{row}
            m_cell = ws.cell(row=row, column=13)
            m_cell.value = f"=G{row}-L{row}"
            _apply_style(m_cell, style[13])

            # N 列: 公式 =G{row}+J{row}+K{row}-L{row}
            n_cell = ws.cell(row=row, column=14)
            n_cell.value = f"=G{row}+J{row}+K{row}-L{row}"
            _apply_style(n_cell, style[14])

            # O 列: 写入 WPS DISPIMG 公式
            o_cell = ws.cell(row=row, column=15)
            _apply_style(o_cell, style[15])

            image_path = record.get("image_path")
            if image_path and os.path.exists(image_path):
                # 使用 UUID 作为占位符，避免不同行之间的占位符冲突
                placeholder = f"PH_{uuid.uuid4().hex.upper()}"
                o_cell.value = f'=_xlfn.DISPIMG("{placeholder}",1)'
                all_image_records.append((row, image_path, placeholder, sheet_idx))
            else:
                o_cell.value = "(无图片)"

            # P 列: 备注
            p_cell = ws.cell(row=row, column=16)
            p_cell.value = record.get("remark", "")
            _apply_style(p_cell, style[16])

            current_row += 1

    # 保存 openpyxl 结果
    _enable_formula_recalculation(wb)
    wb.save(output_path)

    # 注入 WPS 格式的嵌入图片
    if all_image_records:
        # 转换格式：(row, image_path, placeholder) -> 与 _inject_wps_cellimages 兼容
        _inject_wps_cellimages(output_path, [
            (r, img, ph) for r, img, ph, _ in all_image_records
        ])

    return output_path
